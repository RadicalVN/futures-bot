import os
import math
from datetime import datetime, timezone, timedelta
from typing import Optional
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import select
from loguru import logger

from src.database.db import get_db
from src.database.models import Bot, ExchangeAccount
from src.core.exchange import BinanceExchange, create_exchange_from_env
from src.strategies.ma_macd import MaMacdStrategy
from src.strategies.custom_sma import CustomSMAStrategy
from src.strategies.custom_macd import CustomMACDStrategy
from src.strategies.sma_trend_early_exit import SmaTrendEarlyExitStrategy
from src.strategies.sma_pullback import SmaPullbackStrategy
from src.strategies.sma_anti_sideway import SmaAntiSidewayStrategy
from src.strategies.sma_macd_cross import SmaMacdCrossStrategy
from src.strategies.sma_macd_cross_v2 import SmaMacdCrossV2Strategy
from src.strategies.sma_macd_cross_v3 import SmaMacdCrossV3Strategy
from src.strategies.sma_macd_cross_v4 import SmaMacdCrossV4Strategy
from src.strategies.sma_macd_cross_v5 import SmaMacdCrossV5Strategy
from src.strategies.sma_macd_cross_v6 import SmaMacdCrossV6Strategy
from src.strategies.sma_macd_cross_v7 import SmaMacdCrossV7Strategy
from src.strategies.adts import ADTSStrategy
from src.strategies.adts.models import ADTSConfig
from src.strategies.adts.scanner import run_calibration, _calculate_atr, _calculate_bbwidth
from src.strategies.adts.indicators import (
    calculate_adx, calculate_ema, calculate_ema_slope, calculate_atr as _calc_atr_intraday
)
router = APIRouter(prefix='/api/backtest', tags=['Backtest'])
BACKTEST_DIR = 'data/backtest'
# COMMISSION hằng số cũ đã được thay bằng commission_pct động trong parameters
# Giữ lại để tương thích ngược với code cũ nếu cần
COMMISSION = 0.0005
UTC7 = timezone(timedelta(hours=7))


class BacktestRequest(BaseModel):
    bot_id: int
    start_date: str
    end_date: Optional[str] = None
    initial_balance: float = 10000.0
    timeframe: Optional[str] = None       # None = dùng timeframe của bot
    stop_loss_pct: Optional[float] = None  # None = dùng từ params bot
    take_profit_pct: Optional[float] = None


class StrategyBacktestRequest(BaseModel):
    """Chạy backtest theo chiến lược + cặp tiền, không cần bot_id."""
    strategy_name: str           # sma_macd_cross / v2 / v3 / v4 / adts
    symbol: str                  # BTCUSDT, TRUMPUSDT, ...
    start_date: str
    end_date: Optional[str] = None
    initial_balance: float = 10000.0
    # Params chung
    timeframe: Optional[str] = None         # mặc định 5m
    bb_length: Optional[int] = None         # mặc định theo version
    # V2/V3/V4
    use_trend_filter: Optional[bool] = None
    # V3
    min_ma_distance_pct: Optional[float] = None
    min_hold_candles: Optional[int] = None
    # V4
    leverage_v4: Optional[int] = None
    notional_usdt: Optional[float] = None
    stop_loss_pct: Optional[float] = None
    take_profit_pct: Optional[float] = None
    # ADTS params
    adts_atr_period: Optional[int] = None
    adts_adx_period: Optional[int] = None
    adts_ema_period: Optional[int] = None
    adts_ema200_period: Optional[int] = None          # EMA200 Trend Filter period
    adts_bb_period: Optional[int] = None
    adts_bbwidth_sma_period: Optional[int] = None
    adts_adx_threshold: Optional[float] = None
    adts_bbwidth_threshold_factor: Optional[float] = None
    adts_risk_pct: Optional[float] = None
    adts_sl_atr_mult: Optional[float] = None
    adts_hard_sl_pct: Optional[float] = None             # Hard SL % giá entry (mặc định 3%)
    adts_tp1_rr: Optional[float] = None
    adts_tp2_trail_atr_mult: Optional[float] = None
    adts_emergency_adx_threshold: Optional[float] = None
    adts_leverage: Optional[int] = None
    adts_min_notional: Optional[float] = None   # USDT notional tối thiểu sau partial close
    # V6 params
    adx_entry_threshold: Optional[float] = None  # Ngưỡng ADX vào lệnh (env ADX_ENTRY_THRESHOLD, mặc định 25)
    adx_exit_threshold:  Optional[float] = None  # Ngưỡng ADX đóng lệnh (env ADX_EXIT_THRESHOLD, mặc định 25)
    # V7 params
    bb_period: Optional[int]   = None   # Chu kỳ Bollinger Bands (mặc định 20)
    bb_mult:   Optional[float] = None   # Hệ số std BB (mặc định 2.0)
    # Phí giao dịch & trượt giá (áp dụng cho mọi chiến lược)
    commission_pct: Optional[float] = None   # % hoa hồng mỗi lần khớp lệnh (mặc định 0.05%)
    slippage_pct: Optional[float] = None     # % trượt giá mỗi lần khớp lệnh (mặc định 0.0%)


def _timeframe_ms(tf):
    units = {"m": 60000, "h": 3600000, "d": 86400000, "w": 604800000}
    try:
        return int(tf[:-1]) * units[tf[-1]]
    except Exception:
        return 300000


def _build_strategy(strategy_name, parameters):
    if strategy_name == "ma_macd":
        return MaMacdStrategy(parameters)
    elif strategy_name == "custom_sma":
        return CustomSMAStrategy(parameters)
    elif strategy_name == "custom_macd":
        return CustomMACDStrategy(parameters)
    elif strategy_name == "sma_trend_early_exit":
        return SmaTrendEarlyExitStrategy(parameters)
    elif strategy_name == "sma_pullback":
        return SmaPullbackStrategy(parameters)
    elif strategy_name == "sma_anti_sideway":
        return SmaAntiSidewayStrategy(parameters)
    elif strategy_name == "sma_macd_cross":
        return SmaMacdCrossStrategy(parameters)
    elif strategy_name == "sma_macd_cross_v2":
        return SmaMacdCrossV2Strategy(parameters)
    elif strategy_name == "sma_macd_cross_v3":
        return SmaMacdCrossV3Strategy(parameters)
    elif strategy_name == "sma_macd_cross_v4":
        return SmaMacdCrossV4Strategy(parameters)
    elif strategy_name == "sma_macd_cross_v5":
        return SmaMacdCrossV5Strategy(parameters)
    elif strategy_name == "sma_macd_cross_v6":
        return SmaMacdCrossV6Strategy(parameters)
    elif strategy_name == "sma_macd_cross_v7":
        return SmaMacdCrossV7Strategy(parameters)
    elif strategy_name == "adts":
        return ADTSStrategy(parameters)
    else:
        raise ValueError(f"Unsupported strategy: {strategy_name}")


def _get_lookback(strategy_name, parameters):
    base = parameters.get("lookback_candles", 200)
    if strategy_name == "sma_macd_cross":
        # EMA(span=N, adjust=False) cần ~3×N nến để hội tụ đủ (sai số <1%)
        signal_len = int(parameters.get("macd_signal_length", 500))
        return max(base, signal_len * 3)
    elif strategy_name in ("sma_macd_cross_v2", "sma_macd_cross_v3",
                           "sma_macd_cross_v4", "sma_macd_cross_v5"):
        signal_len = int(parameters.get("macd_signal_length", 500))
        return max(base, signal_len * 3)
    elif strategy_name == "sma_macd_cross_v7":
        signal_len = int(parameters.get("macd_signal_length", 500))
        bb_period  = int(parameters.get("bb_period", 20))
        return max(base, signal_len * 3, bb_period)
    elif strategy_name == "sma_macd_cross_v6":
        signal_len = int(parameters.get("macd_signal_length", 500))
        adx_period = int(parameters.get("adx_period", int(float(os.environ.get("ADX_PERIOD", 14)))))
        adx_warmup = adx_period * 7
        return max(base, signal_len * 3, adx_warmup)
    elif strategy_name == "custom_macd":
        signal_len = int(parameters.get("signal_length", 500))
        return max(base, signal_len * 3)
    elif strategy_name == "adts":
        # ADTS intraday chỉ cần warmup cho ATR(14) + ADX(14) + EMA(20) + BBWidth(20)
        return max(base, 300)
    return base


def _normalize_symbol(symbol):
    if "/" in symbol:
        return symbol
    for quote in ("USDT", "BUSD", "BTC", "ETH", "BNB"):
        if symbol.endswith(quote):
            return f"{symbol[:-len(quote)]}/{quote}"
    return symbol


def _to_utc7_str(ts_ms):
    dt = datetime.fromtimestamp(ts_ms / 1000, tz=UTC7)
    return dt.strftime("%Y-%m-%d %H:%M")


async def _fetch_ohlcv_range(exchange, symbol, timeframe, start_ms, end_ms):
    """Fetch OHLCV từ start_ms đến end_ms, paginate ngược nếu cần."""
    all_candles = []
    current_end = end_ms
    max_iterations = 50  # safety cap
    iteration = 0

    while iteration < max_iterations:
        iteration += 1
        batch = await exchange.fetch_ohlcv(
            symbol, timeframe=timeframe, limit=1500,
            params={'endTime': current_end}
        )
        if not batch:
            break

        # Prepend batch (older candles go first)
        all_candles = batch + all_candles
        oldest_ts = batch[0][0]

        if oldest_ts <= start_ms:
            break  # Đã có đủ dữ liệu từ start_ms trở đi

        current_end = oldest_ts - 1
        if current_end < start_ms:
            break

        # Yield control sau mỗi batch để không block event loop
        await _asyncio.sleep(0)

    # Deduplicate và sort
    seen = set()
    result = []
    for c in all_candles:
        if c[0] not in seen:
            seen.add(c[0])
            result.append(c)
    result.sort(key=lambda x: x[0])
    return result


# ── Job store (in-memory) ─────────────────────────────────────────────────────
# job_id → {"status": "running"|"done"|"error", "progress": 0-100,
#            "message": str, "result": dict|None, "error": str|None}
import uuid
import asyncio as _asyncio
_jobs: dict = {}


# ── Pre-compute indicators for sma_macd_cross ────────────────────────────────

def _precompute_sma_macd(df, parameters):
    """
    Tính toàn bộ indicators 1 lần trên DataFrame đầy đủ.
    Trả về DataFrame với các cột indicator đã tính sẵn.
    Nhanh hơn ~100x so với tính lại mỗi nến.
    """
    import pandas as pd
    from src.data.indicators import add_custom_sma_to_df, add_custom_macd_to_df, add_adx_to_df
    from src.strategies.sma_macd_cross import _slope_color, _find_signal_phase_start, SIG_BULLISH, SIG_BEARISH

    df = add_custom_sma_to_df(
        df,
        fast_len=parameters.get("fast_len", 1),
        slow_len=parameters.get("slow_len", 5),
        len_c=parameters.get("len_c", 200),
        factor=parameters.get("factor", 0.05),
        bb_length=parameters.get("bb_length", 50),
    )
    df = add_custom_macd_to_df(
        df,
        fast=parameters.get("macd_fast", 12),
        slow=parameters.get("macd_slow", 26),
        signal_length=parameters.get("macd_signal_length", 500),
        src=parameters.get("macd_src", "EMA"),
        sig_type=parameters.get("macd_sig_type", "EMA"),
    )
    df = add_adx_to_df(df, period=parameters.get("adx_period", int(float(os.environ.get("ADX_PERIOD", 14)))))

    # Pre-compute màu slope cho từng nến
    ma_arr   = df["custom_sma_basis"].to_numpy()
    sig_arr  = df["custom_macd_signal"].to_numpy()
    macd_arr = df["custom_macd"].to_numpy()
    n = len(df)

    ma_colors   = ["yellow"] * n
    sig_colors  = ["yellow"] * n
    macd_colors = ["yellow"] * n

    for i in range(2, n):
        ma_colors[i]   = _slope_color(ma_arr[i],   ma_arr[i-1],   ma_arr[i-2])
        sig_colors[i]  = _slope_color(sig_arr[i],  sig_arr[i-1],  sig_arr[i-2])
        macd_colors[i] = _slope_color(macd_arr[i], macd_arr[i-1], macd_arr[i-2])

    df["_ma_color"]   = ma_colors
    df["_sig_color"]  = sig_colors
    df["_macd_color"] = macd_colors

    # Pre-compute sig_phase_start_ts cho từng nến
    # Duyệt forward: khi sig_color đổi nhóm → phase mới bắt đầu
    phase_starts = [0] * n
    current_phase_start = 0
    current_phase_group = None  # "bullish" | "bearish" | None

    for i in range(n):
        sc = sig_colors[i]
        if sc in SIG_BULLISH:
            group = "bullish"
        elif sc in SIG_BEARISH:
            group = "bearish"
        else:
            group = None

        if group != current_phase_group:
            current_phase_group = group
            current_phase_start = i

        phase_starts[i] = current_phase_start

    df["_sig_phase_start_idx"] = phase_starts

    return df


def _simulate_sma_macd_candle(df, i, open_position, last_entry_phase, parameters, strategy_name="sma_macd_cross"):
    """
    Simulate 1 nến cho sma_macd_cross / sma_macd_cross_v2 dùng pre-computed indicators.
    Trả về signal dict: {"type": "long"|"short"|"close_long"|"close_short"|"none", "price": float, "metadata": dict}
    """
    from src.strategies.sma_macd_cross import SIG_BULLISH, SIG_BEARISH

    row = df.iloc[i]
    close_curr = float(row["close"])
    close_prev = float(df["close"].iloc[i-1])
    high_curr  = float(row["high"])
    low_curr   = float(row["low"])

    ma_curr  = float(row["custom_sma_basis"])
    ma_prev  = float(df["custom_sma_basis"].iloc[i-1])
    macd_curr = float(row["custom_macd"])
    sig_curr  = float(row["custom_macd_signal"])

    ma_color   = row["_ma_color"]
    sig_color  = row["_sig_color"]
    macd_color = row["_macd_color"]

    phase_start_idx = int(row["_sig_phase_start_idx"])
    sig_phase_start_ts = int(df["timestamp"].iloc[phase_start_idx])

    # Trend từ custom_sma (dùng cho V2 trend filter)
    trend_curr = int(row.get("custom_sma_trend", 0))
    use_trend_filter = parameters.get("use_trend_filter", False)  # V1: False, V2: True (set trong strategy)

    # V3 params
    min_ma_distance_pct = float(parameters.get("min_ma_distance_pct", 0.0))
    min_hold_candles    = int(parameters.get("min_hold_candles", 0))
    # V4 params
    stop_loss_pct   = float(parameters.get("stop_loss_pct", 0.0))
    take_profit_pct = float(parameters.get("take_profit_pct", 0.0))
    # V6 params — tách biệt entry / exit threshold
    import os as _os
    _adx_period_default    = int(float(_os.environ.get("ADX_PERIOD", 14)))
    _adx_entry_default     = float(_os.environ.get("ADX_ENTRY_THRESHOLD", 25.0))
    _adx_exit_default      = float(_os.environ.get("ADX_EXIT_THRESHOLD",  25.0))
    adx_entry_threshold    = float(parameters.get("adx_entry_threshold", _adx_entry_default))
    adx_exit_threshold     = float(parameters.get("adx_exit_threshold",  _adx_exit_default))
    adx_curr               = float(row.get("adx", 0) or 0)

    # Tính số nến đã giữ lệnh (cho V3 min_hold)
    candles_held = 0
    if open_position and min_hold_candles > 0:
        entry_candle_ts = int((open_position.get("metadata") or {}).get("entry_candle_ts", 0) or 0)
        if entry_candle_ts:
            ts_arr = [int(t) for t in df["timestamp"].tolist()]
            try:
                entry_idx = next(idx for idx, t in enumerate(ts_arr) if t >= entry_candle_ts)
                candles_held = i - entry_idx
            except StopIteration:
                candles_held = 0

    # ── EXIT ──────────────────────────────────────────────────────────────────
    if open_position:
        side = open_position["side"]
        pos_ma_cross = float((open_position.get("metadata") or {}).get("ma_cross_price", open_position["entry_price"]))
        pos_dev      = float((open_position.get("metadata") or {}).get("entry_deviation", 0))
        # V4: lay entry_price tu metadata (chinh xac hon)
        pos_ep = float((open_position.get("metadata") or {}).get("entry_price", 0) or open_position.get("entry_price", 0) or 0)

        if side == "long":
            # V4: chi SL/TP theo % notional, bo TH1/TH2/TH3
            if stop_loss_pct > 0 and pos_ep > 0:
                notional = float(parameters.get("notional_usdt", 0)) or (float(parameters.get("position_size_pct", 0.1)) * 10000 * float(parameters.get("leverage", 5)))
                sl = pos_ep * (1 - stop_loss_pct / 100)
                tp = pos_ep * (1 + take_profit_pct / 100) if take_profit_pct > 0 else None
                if low_curr <= sl:
                    sl_exit = min(sl, close_curr)
                    return {"type": "close_long", "price": sl_exit, "reason": f"SL: low={low_curr:.4f}<=SL={sl:.4f} (-{stop_loss_pct}%=${notional*stop_loss_pct/100:.2f})"}
                if tp and high_curr >= tp:
                    tp_exit = max(tp, close_curr)
                    return {"type": "close_long", "price": tp_exit, "reason": f"TP: high={high_curr:.4f}>=TP={tp:.4f} (+{take_profit_pct}%=${notional*take_profit_pct/100:.2f})"}
                return {"type": "none", "price": close_curr}

            # V6: kiểm tra ADX trước khi xét TH1/TH2
            if strategy_name == "sma_macd_cross_v6" and adx_curr <= adx_exit_threshold:
                return {"type": "none", "price": close_curr,
                        "reason": f"Giữ LONG: ADX={adx_curr:.2f}≤{adx_exit_threshold}"}

            # V1/V2/V3/V6: TH2/TH1
            if sig_color in SIG_BEARISH:
                # V6 TH2: thêm điều kiện close < ma_curr
                if strategy_name == "sma_macd_cross_v6" and close_curr >= ma_curr:
                    pass  # giá chưa dưới MA → chưa đóng TH2
                else:
                    return {"type": "close_long", "price": close_curr, "reason": f"TH2: Signal {sig_color}"}
            if close_curr < ma_curr:
                if candles_held >= min_hold_candles:
                    threshold = pos_ma_cross + pos_dev
                    if close_curr < threshold:
                        exit_price = (low_curr + ma_curr) / 2
                        return {"type": "close_long", "price": exit_price, "reason": f"TH1: close<MA hold={candles_held}"}

        elif side == "short":
            # V4: chi SL/TP theo % notional, bo TH1/TH2/TH3
            if stop_loss_pct > 0 and pos_ep > 0:
                notional = float(parameters.get("notional_usdt", 0)) or (float(parameters.get("position_size_pct", 0.1)) * 10000 * float(parameters.get("leverage", 5)))
                sl = pos_ep * (1 + stop_loss_pct / 100)
                tp = pos_ep * (1 - take_profit_pct / 100) if take_profit_pct > 0 else None
                if high_curr >= sl:
                    sl_exit = max(sl, close_curr)
                    return {"type": "close_short", "price": sl_exit, "reason": f"SL: high={high_curr:.4f}>=SL={sl:.4f} (-{stop_loss_pct}%=${notional*stop_loss_pct/100:.2f})"}
                if tp and low_curr <= tp:
                    tp_exit = min(tp, close_curr)
                    return {"type": "close_short", "price": tp_exit, "reason": f"TP: low={low_curr:.4f}<=TP={tp:.4f} (+{take_profit_pct}%=${notional*take_profit_pct/100:.2f})"}
                return {"type": "none", "price": close_curr}

            # V6: kiểm tra ADX trước khi xét TH1/TH2
            if strategy_name == "sma_macd_cross_v6" and adx_curr <= adx_exit_threshold:
                return {"type": "none", "price": close_curr,
                        "reason": f"Giữ SHORT: ADX={adx_curr:.2f}≤{adx_exit_threshold}"}

            # V1/V2/V3/V6: TH2/TH1
            if sig_color in SIG_BULLISH:
                # V6 TH2: thêm điều kiện close > ma_curr
                if strategy_name == "sma_macd_cross_v6" and close_curr <= ma_curr:
                    pass  # giá chưa trên MA → chưa đóng TH2
                else:
                    return {"type": "close_short", "price": close_curr, "reason": f"TH2: Signal {sig_color}"}
            if close_curr > ma_curr:
                if candles_held >= min_hold_candles:
                    threshold = pos_ma_cross + pos_dev
                    if close_curr > threshold:
                        exit_price = (high_curr + ma_curr) / 2
                        return {"type": "close_short", "price": exit_price, "reason": f"TH1: close>MA hold={candles_held}"}

        return {"type": "none", "price": close_curr}

    # ── ENTRY ─────────────────────────────────────────────────────────────────
    # Khoang cach gia-MA (V3)
    ma_dist_pct = abs(close_curr - ma_curr) / ma_curr * 100 if ma_curr else 0

    # LONG
    cond1_long = sig_color in SIG_BULLISH
    cond2_long = macd_curr >= sig_curr
    cond3_long = (close_prev <= ma_prev) and (close_curr > ma_curr)
    cond4_long = (not use_trend_filter) or (trend_curr == 1)
    cond5_long = ma_dist_pct >= min_ma_distance_pct  # V3: khoang cach toi thieu
    # V5: MA200 phai di ngang hoac doc len
    ma_long_ok  = {"blue", "green", "yellow"}
    cond6_long  = (strategy_name != "sma_macd_cross_v5") or (ma_color in ma_long_ok)
    # V6: ADX > entry threshold
    cond7_long  = (strategy_name != "sma_macd_cross_v6") or (adx_curr > adx_entry_threshold)

    if cond1_long and cond2_long and cond3_long and cond4_long and cond5_long and cond6_long and cond7_long:
        # One-shot check
        if "long" in last_entry_phase and last_entry_phase["long"] == sig_phase_start_ts:
            return {"type": "none", "price": close_curr}
        ma_cross = ma_curr
        entry_price = (high_curr + ma_cross) / 2
        deviation = abs(entry_price - ma_cross)
        curr_ts = int(df["timestamp"].iloc[i])
        return {
            "type": "long", "price": entry_price,
            "metadata": {
                "entry_price": round(entry_price, 6),   # quan trong cho V4 SL/TP
                "ma_cross_price": round(ma_cross, 6),
                "entry_deviation": round(deviation, 6),
                "entry_candle_ts": curr_ts,
                "sig_phase_start_ts": sig_phase_start_ts,
                "ma_color": ma_color, "sig_color": sig_color, "macd_color": macd_color,
                "ma": round(ma_curr, 6), "close": round(close_curr, 6),
            }
        }

    # SHORT
    cond1_short = sig_color in SIG_BEARISH
    cond2_short = macd_curr <= sig_curr
    cond3_short = (close_prev >= ma_prev) and (close_curr < ma_curr)
    cond4_short = (not use_trend_filter) or (trend_curr == -1)
    cond5_short = ma_dist_pct >= min_ma_distance_pct  # V3: khoang cach toi thieu
    # V5: MA200 phai di ngang hoac doc xuong
    ma_short_ok  = {"red", "orange", "yellow"}
    cond6_short  = (strategy_name != "sma_macd_cross_v5") or (ma_color in ma_short_ok)
    # V6: ADX > entry threshold
    cond7_short  = (strategy_name != "sma_macd_cross_v6") or (adx_curr > adx_entry_threshold)

    if cond1_short and cond2_short and cond3_short and cond4_short and cond5_short and cond6_short and cond7_short:
        # One-shot check
        if "short" in last_entry_phase and last_entry_phase["short"] == sig_phase_start_ts:
            return {"type": "none", "price": close_curr}
        ma_cross = ma_curr
        entry_price = (low_curr + ma_cross) / 2
        deviation = abs(entry_price - ma_cross)
        curr_ts = int(df["timestamp"].iloc[i])
        return {
            "type": "short", "price": entry_price,
            "metadata": {
                "entry_price": round(entry_price, 6),   # quan trong cho V4 SL/TP
                "ma_cross_price": round(ma_cross, 6),
                "entry_deviation": round(deviation, 6),
                "entry_candle_ts": curr_ts,
                "sig_phase_start_ts": sig_phase_start_ts,
                "ma_color": ma_color, "sig_color": sig_color, "macd_color": macd_color,
                "ma": round(ma_curr, 6), "close": round(close_curr, 6),
            }
        }

    return {"type": "none", "price": close_curr}


# ── Simulator for sma_macd_cross_v7 ─────────────────────────────────────────

def _simulate_sma_macd_v7_candle(df, i, open_position, last_entry_phase, parameters):
    """
    Simulate 1 nến cho sma_macd_cross_v7 dùng pre-computed indicators.

    Thay đổi so với V1:
    - Entry cond1: sig_curr > 0 (LONG) / sig_curr < 0 (SHORT)
    - Exit: chỉ TH1, không có TH2
    - TH1 LONG : close < MA AND close < threshold AND close > bb_upper
    - TH1 SHORT: close > MA AND close > threshold AND close < bb_lower
    """
    from src.strategies.sma_macd_cross import SIG_BULLISH, SIG_BEARISH

    row        = df.iloc[i]
    close_curr = float(row["close"])
    close_prev = float(df["close"].iloc[i - 1])
    high_curr  = float(row["high"])
    low_curr   = float(row["low"])

    ma_curr    = float(row["custom_sma_basis"])
    ma_prev    = float(df["custom_sma_basis"].iloc[i - 1])
    macd_curr  = float(row["custom_macd"])
    sig_curr   = float(row["custom_macd_signal"])

    ma_color   = row["_ma_color"]
    sig_color  = row["_sig_color"]
    macd_color = row["_macd_color"]

    phase_start_idx    = int(row["_sig_phase_start_idx"])
    sig_phase_start_ts = int(df["timestamp"].iloc[phase_start_idx])

    bb_upper = float(row.get("bb_upper", float("nan")))
    bb_lower = float(row.get("bb_lower", float("nan")))

    import math
    bb_valid = not (math.isnan(bb_upper) or math.isnan(bb_lower))

    # TVT-MA color: dung custom_sma_momentum (tinh stateful, nhat quan voi bieu do)
    # Khac voi _ma_color (tinh tu _slope_color, phu thuoc so nen load)
    tvt_ma_color = str(row.get("custom_sma_momentum", "yellow") or "yellow")

    # ── EXIT ──────────────────────────────────────────────────────────────────
    if open_position:
        side         = open_position["side"]
        pos_ma_cross = float((open_position.get("metadata") or {}).get("ma_cross_price", open_position["entry_price"]))
        pos_dev      = float((open_position.get("metadata") or {}).get("entry_deviation", 0))

        if side == "long":
            # TH1: close < MA AND close < threshold AND close > bb_upper
            #      AND sig_color in SIG_BEARISH
            if close_curr < ma_curr and bb_valid:
                threshold = pos_ma_cross + pos_dev
                if (close_curr < threshold
                        and close_curr > bb_upper
                        and sig_color in SIG_BEARISH):
                    exit_price = (low_curr + ma_curr) / 2
                    return {
                        "type": "close_long", "price": exit_price,
                        "reason": (
                            f"TH1: close<MA, <thr, >BB_Upper={bb_upper:.4f}"
                            f", Signal={sig_color}(bearish)"
                        ),
                    }

        elif side == "short":
            # TH1: close > MA AND close > threshold AND close < bb_lower
            #      AND sig_color in SIG_BULLISH
            if close_curr > ma_curr and bb_valid:
                threshold = pos_ma_cross + pos_dev
                if (close_curr > threshold
                        and close_curr < bb_lower
                        and sig_color in SIG_BULLISH):
                    exit_price = (high_curr + ma_curr) / 2
                    return {
                        "type": "close_short", "price": exit_price,
                        "reason": (
                            f"TH1: close>MA, >thr, <BB_Lower={bb_lower:.4f}"
                            f", Signal={sig_color}(bullish)"
                        ),
                    }

        return {"type": "none", "price": close_curr}

    # ── ENTRY ─────────────────────────────────────────────────────────────────
    # LONG: Signal > 0, MACD >= Signal, giá cắt lên MA, TVT-SMA xanh
    cond1_long = sig_curr > 0
    cond2_long = macd_curr >= sig_curr
    cond3_long = (close_prev <= ma_prev) and (close_curr > ma_curr)
    cond4_long = tvt_ma_color in {"blue", "green"}  # TVT-MA xanh (stateful, nhat quan voi bieu do)

    if cond1_long and cond2_long and cond3_long and cond4_long:
        if "long" in last_entry_phase and last_entry_phase["long"] == sig_phase_start_ts:
            return {"type": "none", "price": close_curr}
        ma_cross    = ma_curr
        entry_price = (high_curr + ma_cross) / 2
        deviation   = abs(entry_price - ma_cross)
        curr_ts     = int(df["timestamp"].iloc[i])
        return {
            "type": "long", "price": entry_price,
            "metadata": {
                "entry_price":       round(entry_price, 6),
                "ma_cross_price":    round(ma_cross, 6),
                "entry_deviation":   round(deviation, 6),
                "entry_candle_ts":   curr_ts,
                "sig_phase_start_ts": sig_phase_start_ts,
                "ma_color": ma_color, "sig_color": sig_color, "macd_color": macd_color,
                "ma": round(ma_curr, 6), "close": round(close_curr, 6),
                "bb_upper": round(bb_upper, 6) if bb_valid else None,
                "bb_lower": round(bb_lower, 6) if bb_valid else None,
            },
        }

    # SHORT: Signal < 0, MACD <= Signal, giá cắt xuống MA, TVT-SMA đỏ/cam
    cond1_short = sig_curr < 0
    cond2_short = macd_curr <= sig_curr
    cond3_short = (close_prev >= ma_prev) and (close_curr < ma_curr)
    cond4_short = tvt_ma_color in {"red", "orange"}  # TVT-MA do/cam (stateful, nhat quan voi bieu do)

    if cond1_short and cond2_short and cond3_short and cond4_short:
        if "short" in last_entry_phase and last_entry_phase["short"] == sig_phase_start_ts:
            return {"type": "none", "price": close_curr}
        ma_cross    = ma_curr
        entry_price = (low_curr + ma_cross) / 2
        deviation   = abs(entry_price - ma_cross)
        curr_ts     = int(df["timestamp"].iloc[i])
        return {
            "type": "short", "price": entry_price,
            "metadata": {
                "entry_price":       round(entry_price, 6),
                "ma_cross_price":    round(ma_cross, 6),
                "entry_deviation":   round(deviation, 6),
                "entry_candle_ts":   curr_ts,
                "sig_phase_start_ts": sig_phase_start_ts,
                "ma_color": ma_color, "sig_color": sig_color, "macd_color": macd_color,
                "ma": round(ma_curr, 6), "close": round(close_curr, 6),
                "bb_upper": round(bb_upper, 6) if bb_valid else None,
                "bb_lower": round(bb_lower, 6) if bb_valid else None,
            },
        }

    return {"type": "none", "price": close_curr}


# ── Pre-compute indicators for ADTS ──────────────────────────────────────────

def _precompute_adts(df, parameters, d1_candles=None):
    """
    Tính toàn bộ indicators ADTS 1 lần trên DataFrame đầy đủ.
    Bao gồm:
      - ATR(14) intraday
      - ADX(14) intraday
      - BBWidth(20,2) intraday
      - EMA20 + EMA20_Slope intraday
      - Daily calibration: Base_ATR_D1, Sideway_Threshold, Min_Slope
        → Tính lại mỗi ngày (simulate đúng hành vi production)

    Args:
        df: DataFrame intraday OHLCV
        parameters: dict tham số ADTS
        d1_candles: list [[ts_ms, o, h, l, c, v], ...] dữ liệu D1 đã fetch riêng.
                    Nếu None → fallback resample từ intraday (chỉ dùng khi test).

    Returns DataFrame với các cột indicator đã tính sẵn.
    """
    import pandas as pd
    import numpy as np

    cfg = ADTSConfig.from_dict(parameters)

    # ── Intraday indicators ───────────────────────────────────────────────────
    df = df.copy()
    df["_atr"]         = _calc_atr_intraday(df, cfg.atr_period)
    df["_adx"]         = calculate_adx(df, cfg.adx_period)
    df["_ema20"]       = calculate_ema(df["close"], cfg.ema_period)
    df["_ema20_slope"] = calculate_ema_slope(df, cfg.ema_period)
    df["_ema200"]      = calculate_ema(df["close"], cfg.ema200_period)

    # BBWidth intraday
    close  = df["close"]
    bb_mid = close.rolling(cfg.bb_period).mean()
    bb_std = close.rolling(cfg.bb_period).std(ddof=0)
    df["_bbwidth"] = (bb_mid + 2 * bb_std - (bb_mid - 2 * bb_std)) / bb_mid

    # ── Daily calibration ─────────────────────────────────────────────────────
    df["_date"] = pd.to_datetime(df["timestamp"], unit="ms", utc=True).dt.date

    if d1_candles and len(d1_candles) >= cfg.atr_period + 2:
        # Dùng D1 data đã fetch riêng — chính xác và đủ lịch sử
        d1 = pd.DataFrame(
            d1_candles,
            columns=["timestamp", "open", "high", "low", "close", "volume"],
        ).astype({"open": float, "high": float, "low": float, "close": float, "volume": float})
        d1["_dt"] = pd.to_datetime(d1["timestamp"], unit="ms", utc=True)
        d1 = d1.set_index("_dt")
    else:
        # Fallback: resample từ intraday (chỉ dùng khi test với dữ liệu giả)
        df_ts = df.copy()
        df_ts["_ts_dt"] = pd.to_datetime(df_ts["timestamp"], unit="ms", utc=True)
        df_ts = df_ts.set_index("_ts_dt")
        d1 = df_ts[["open", "high", "low", "close", "volume"]].resample("1D").agg({
            "open": "first", "high": "max", "low": "min",
            "close": "last", "volume": "sum",
        }).dropna()

    # Tính ATR D1 và BBWidth D1
    d1_atr_series = _calc_atr_intraday(d1, cfg.atr_period)
    d1_bbw_series = _calculate_bbwidth(d1, cfg.bb_period, cfg.bb_std)
    d1_bbw_sma    = d1_bbw_series.rolling(cfg.bbwidth_sma_period).mean()

    # shift(1): dùng giá trị ngày hôm trước để tránh look-ahead bias
    d1_calib = pd.DataFrame({
        "base_atr":    d1_atr_series.shift(1),
        "bbwidth_sma": d1_bbw_sma.shift(1),
    }, index=d1.index)
    d1_calib["sideway_threshold"] = d1_calib["bbwidth_sma"] * cfg.bbwidth_threshold_factor
    d1_calib["min_slope"]         = (d1_calib["base_atr"] * cfg.min_slope_atr_factor) / 5.0
    d1_calib.index = d1_calib.index.date

    # Map về intraday theo ngày
    # Dùng reindex + ffill thay vì map() để xử lý đúng khi intraday date
    # không có trong d1_calib (ví dụ: ngày cuối cùng của D1 chưa đóng nến)
    d1_calib_dated = d1_calib.copy()
    d1_calib_dated.index = pd.to_datetime(d1_calib_dated.index)

    # Tạo DatetimeIndex cho từng ngày intraday
    intraday_dates = pd.to_datetime(df["_date"])

    # Reindex D1 calibration về intraday dates, forward-fill
    calib_reindexed = d1_calib_dated.reindex(
        intraday_dates, method="ffill"
    )

    df["_calib_base_atr"]  = calib_reindexed["base_atr"].values
    df["_calib_sideway"]   = calib_reindexed["sideway_threshold"].values
    df["_calib_min_slope"] = calib_reindexed["min_slope"].values

    return df


def _simulate_adts_candle(df, i, open_position, parameters, _cfg=None):
    """
    Simulate 1 nến cho ADTS dùng pre-computed indicators.
    Trả về signal dict: {"type": str, "price": float, "metadata": dict, "partial": bool, "partial_pct": float}

    Xử lý đầy đủ:
      - The Shield (3 điều kiện)
      - Entry: close vs EMA20 + slope direction
      - Exit: SL → TP1 (50%) → TP2 Trailing → Emergency Exit

    Args:
        _cfg: ADTSConfig pre-built (optional). Nếu None sẽ tự build từ parameters.
              Truyền vào để tránh rebuild Pydantic model mỗi nến (tối ưu grid search).
    """
    cfg = _cfg if _cfg is not None else ADTSConfig.from_dict(parameters)
    row = df.iloc[i]

    close  = float(row["close"])
    high   = float(row["high"])
    low    = float(row["low"])
    atr    = float(row["_atr"])
    adx    = float(row["_adx"])
    bbw    = float(row["_bbwidth"])
    ema20  = float(row["_ema20"])
    slope  = float(row["_ema20_slope"])
    ema200 = float(row["_ema200"]) if "_ema200" in df.columns else float("nan")

    # Calibration values cho ngày này
    calib_sideway   = row["_calib_sideway"]
    calib_min_slope = row["_calib_min_slope"]

    # Kiểm tra NaN (chưa đủ warmup)
    import math
    if any(math.isnan(v) for v in [atr, adx, bbw, ema20, slope] if isinstance(v, float)):
        return {"type": "none", "price": close}
    if isinstance(calib_sideway, float) and math.isnan(calib_sideway):
        return {"type": "none", "price": close}
    # EMA200 có thể NaN ở những nến đầu — nếu NaN thì bỏ qua trend filter
    ema200_valid = not (isinstance(ema200, float) and math.isnan(ema200))

    # ── The Shield ────────────────────────────────────────────────────────────
    adx_ok    = adx > cfg.adx_threshold
    bbw_ok    = bbw > calib_sideway
    slope_ok  = abs(slope) > calib_min_slope
    shield_ok = adx_ok and bbw_ok and slope_ok

    # ── Trend Filter: EMA200 ──────────────────────────────────────────────────
    above_ema200 = (close > ema200) if ema200_valid else True   # fallback: không chặn
    below_ema200 = (close < ema200) if ema200_valid else True

    meta_base = {
        "adx": round(adx, 2), "bb_width": round(bbw, 6),
        "ema20": round(ema20, 6), "ema20_slope": round(slope, 8),
        "ema200": round(ema200, 6) if ema200_valid else None,
        "above_ema200": above_ema200,
        "atr": round(atr, 6),
        "sideway_threshold": round(float(calib_sideway), 6),
        "min_slope": round(float(calib_min_slope), 8),
        "shield_passed": shield_ok,
        "adx_ok": adx_ok, "bbwidth_ok": bbw_ok, "slope_ok": slope_ok,
        "emergency_adx_threshold": cfg.emergency_adx_threshold,
    }

    # ── EXIT checks (ưu tiên trước entry) ────────────────────────────────────
    if open_position:
        side      = open_position["side"]
        ep        = float(open_position.get("entry_price", 0))
        sl        = float(open_position.get("stop_loss", 0))
        tp1       = float(open_position.get("take_profit_1", 0))
        tp1_hit   = open_position.get("tp1_hit", False)
        trail     = float(open_position.get("trailing_stop", 0))
        atr_entry = float(open_position.get("atr_at_entry", atr))
        # Flags chống lặp lại
        is_emergency_closed = open_position.get("is_emergency_closed", False)

        # 1. Emergency Exit — chỉ kích hoạt MỘT LẦN duy nhất
        if not is_emergency_closed:
            emg = (adx < cfg.emergency_adx_threshold) or (bbw < calib_sideway)
            if emg:
                reason = (
                    f"Emergency: ADX={adx:.1f}<{cfg.emergency_adx_threshold}"
                    if adx < cfg.emergency_adx_threshold
                    else f"Emergency: BBW={bbw:.5f}<Thr={calib_sideway:.5f}"
                )
                # Đánh dấu đã kích hoạt — không cho phép lặp lại ở nến tiếp theo
                open_position["is_emergency_closed"] = True
                return {
                    "type": f"close_{side}", "price": close,
                    "partial": True, "partial_pct": cfg.emergency_close_pct,
                    "reason": reason, "metadata": meta_base,
                }

        # 2. Stop Loss
        if side == "long" and low <= sl:
            return {"type": "close_long", "price": min(sl, close),
                    "partial": False, "reason": f"SL: low={low:.4f}<=SL={sl:.4f}", "metadata": meta_base}
        if side == "short" and high >= sl:
            return {"type": "close_short", "price": max(sl, close),
                    "partial": False, "reason": f"SL: high={high:.4f}>=SL={sl:.4f}", "metadata": meta_base}

        # 3. TP1 — chỉ kích hoạt MỘT LẦN duy nhất (flag tp1_hit)
        if not tp1_hit:
            tp1_triggered = (side == "long" and high >= tp1) or (side == "short" and low <= tp1)
            if tp1_triggered:
                tp1_price = tp1
                # Đánh dấu TP1 đã hit — không cho phép lặp lại ở nến tiếp theo
                open_position["tp1_hit"] = True
                open_position["stop_loss"] = ep          # SL → Entry (break-even)
                open_position["trailing_stop"] = tp1_price  # trailing bắt đầu từ TP1
                hit_val = high if side == "long" else low
                return {
                    "type": f"close_{side}", "price": tp1_price,
                    "partial": True, "partial_pct": cfg.tp1_close_pct,
                    "reason": (
                        f"TP1: {'high' if side=='long' else 'low'}={hit_val:.4f} "
                        f"{'≥' if side=='long' else '≤'} TP1={tp1:.4f} "
                        f"(chốt {cfg.tp1_close_pct*100:.0f}%)"
                    ),
                    "metadata": meta_base,
                }

        # 4. TP2 Trailing Stop (sau khi TP1 hit)
        if tp1_hit:
            trail_dist = cfg.tp2_trail_atr_mult * atr
            if side == "long":
                new_trail = close - trail_dist
                trail = max(trail, new_trail)
                open_position["trailing_stop"] = trail
                if low <= trail:
                    return {"type": "close_long", "price": trail,
                            "partial": False, "reason": f"TP2 Trail: low={low:.4f}<=Trail={trail:.4f}", "metadata": meta_base}
            else:
                new_trail = close + trail_dist
                trail = min(trail, new_trail) if trail > 0 else new_trail
                open_position["trailing_stop"] = trail
                if high >= trail:
                    return {"type": "close_short", "price": trail,
                            "partial": False, "reason": f"TP2 Trail: high={high:.4f}>=Trail={trail:.4f}", "metadata": meta_base}

        return {"type": "none", "price": close}

    # ── ENTRY ─────────────────────────────────────────────────────────────────
    if not shield_ok:
        return {"type": "none", "price": close}

    sl_dist  = cfg.sl_atr_mult * atr
    tp1_dist = sl_dist * cfg.tp1_rr

    # BUY: close > EMA20 + slope > 0 + close > EMA200 (Trend Filter)
    if close > ema20 and slope > 0 and above_ema200:
        atr_sl   = cfg.sl_atr_mult * atr
        hard_sl  = close * cfg.hard_sl_pct
        sl_dist  = min(atr_sl, hard_sl)
        tp1_dist = sl_dist * cfg.tp1_rr
        sl_price   = close - sl_dist
        tp1_price  = close + tp1_dist
        trail_init = close - cfg.tp2_trail_atr_mult * atr
        sl_src = "ATR" if atr_sl <= hard_sl else "Hard"
        return {
            "type": "long", "price": close,
            "stop_loss": sl_price, "take_profit_1": tp1_price,
            "trailing_stop_init": trail_init, "atr_at_entry": atr,
            "metadata": {**meta_base, "entry_price": round(close, 6),
                         "stop_loss": round(sl_price, 6), "take_profit_1": round(tp1_price, 6),
                         "sl_source": sl_src, "sl_distance": round(sl_dist, 6)},
        }

    # SELL: close < EMA20 + slope < 0 + close < EMA200 (Trend Filter)
    if close < ema20 and slope < 0 and below_ema200:
        atr_sl   = cfg.sl_atr_mult * atr
        hard_sl  = close * cfg.hard_sl_pct
        sl_dist  = min(atr_sl, hard_sl)
        tp1_dist = sl_dist * cfg.tp1_rr
        sl_price   = close + sl_dist
        tp1_price  = close - tp1_dist
        trail_init = close + cfg.tp2_trail_atr_mult * atr
        sl_src = "ATR" if atr_sl <= hard_sl else "Hard"
        return {
            "type": "short", "price": close,
            "stop_loss": sl_price, "take_profit_1": tp1_price,
            "trailing_stop_init": trail_init, "atr_at_entry": atr,
            "metadata": {**meta_base, "entry_price": round(close, 6),
                         "stop_loss": round(sl_price, 6), "take_profit_1": round(tp1_price, 6),
                         "sl_source": sl_src, "sl_distance": round(sl_dist, 6)},
        }

    return {"type": "none", "price": close}




# ── Backtest Integrity Validator ──────────────────────────────────────────────

def validate_backtest_integrity(
    trades: list[dict],
    initial_balance: float,
) -> dict:
    """
    Kiểm tra tính toàn vẹn của danh sách lệnh sau khi simulation kết thúc.

    Phát hiện các race condition và lỗi logic:
      1. CLOSE_WITHOUT_OPEN  — lệnh đóng khi không có lệnh nào đang mở
      2. OPEN_WHILE_OPEN     — lệnh mở khi đã có lệnh đang mở (chưa đóng)
      3. SIDE_MISMATCH       — đóng sai chiều (close_long khi đang short, v.v.)
      4. NEGATIVE_BALANCE    — số dư âm sau một lệnh
      5. ZERO_SIZE           — lệnh có size = 0
      6. NEGATIVE_SIZE       — lệnh có size âm
      7. EXIT_BEFORE_ENTRY   — thời gian đóng lệnh trước thời gian mở lệnh
      8. DUPLICATE_TIMESTAMP — hai lệnh cùng entry_ts_ms (entry trùng nến)
      9. BALANCE_DRIFT       — số dư cuối không khớp với initial_balance + sum(pnl)
                               (sai số > 0.01 USDT — dấu hiệu lỗi tính toán)

    Args:
        trades: Danh sách trade records từ simulation loop
        initial_balance: Vốn ban đầu

    Returns:
        {
          "ok": bool,                  # True nếu không có vi phạm
          "violation_count": int,
          "violations": [              # Danh sách vi phạm chi tiết
            {
              "index": int,            # Vị trí trong danh sách trades (0-based)
              "trade_num": int,        # Số thứ tự lệnh (1-based)
              "rule": str,             # Mã lỗi
              "severity": str,         # "error" | "warning"
              "message": str,          # Mô tả chi tiết
              "trade_snapshot": dict,  # Snapshot của trade bị lỗi
            }
          ],
          "stats": {                   # Thống kê tổng hợp
            "total_trades": int,
            "error_count": int,
            "warning_count": int,
            "balance_drift": float,
          }
        }
    """
    violations: list[dict] = []

    def _add(idx: int, rule: str, severity: str, message: str, trade: dict) -> None:
        violations.append({
            "index":          idx,
            "trade_num":      idx + 1,
            "rule":           rule,
            "severity":       severity,
            "message":        message,
            "trade_snapshot": {
                "entry_time":   trade.get("entry_time", ""),
                "exit_time":    trade.get("exit_time", ""),
                "side":         trade.get("side", ""),
                "entry_price":  trade.get("entry_price", 0),
                "exit_price":   trade.get("exit_price", 0),
                "size":         trade.get("size", 0),
                "pnl":          trade.get("pnl", 0),
                "balance_after": trade.get("balance_after", 0),
            },
        })

    # ── Trạng thái giả lập để replay ─────────────────────────────────────────
    # Replay lại chuỗi lệnh để phát hiện vi phạm thứ tự.
    # Mỗi trade record đại diện cho 1 lần đóng lệnh (full hoặc partial).
    # Với partial close: nhiều trade record có cùng entry_ts_ms.
    #
    # open_positions: entry_ts_ms → {"side": str, "last_exit_ts": int}
    #   last_exit_ts = exit_ts của lần đóng gần nhất (để phát hiện overlap)
    open_positions: dict[int, dict] = {}
    seen_entry_ts: set[int] = set()

    for idx, trade in enumerate(trades):
        side        = trade.get("side", "")
        entry_ts    = trade.get("entry_ts_ms", 0)
        exit_ts     = trade.get("exit_ts_ms", 0)
        size        = trade.get("size", 0)
        balance_aft = trade.get("balance_after", 0)
        entry_price = trade.get("entry_price", 0)
        exit_price  = trade.get("exit_price", 0)

        # ── Rule 5: ZERO_SIZE ─────────────────────────────────────────────────
        if size == 0:
            _add(idx, "ZERO_SIZE", "error",
                 f"Lệnh #{idx+1} có size = 0 (entry={trade.get('entry_time')})",
                 trade)

        # ── Rule 6: NEGATIVE_SIZE ─────────────────────────────────────────────
        elif size < 0:
            _add(idx, "NEGATIVE_SIZE", "error",
                 f"Lệnh #{idx+1} có size âm = {size} (entry={trade.get('entry_time')})",
                 trade)

        # ── Rule 7: EXIT_BEFORE_ENTRY ─────────────────────────────────────────
        if exit_ts and entry_ts and exit_ts < entry_ts:
            _add(idx, "EXIT_BEFORE_ENTRY", "error",
                 f"Lệnh #{idx+1}: exit_ts ({trade.get('exit_time')}) < entry_ts ({trade.get('entry_time')})",
                 trade)

        # ── Rule 4: NEGATIVE_BALANCE ──────────────────────────────────────────
        if balance_aft < 0:
            _add(idx, "NEGATIVE_BALANCE", "error",
                 f"Lệnh #{idx+1}: balance_after = {balance_aft:.4f} USDT (âm) sau khi đóng",
                 trade)

        # ── Replay state machine ──────────────────────────────────────────────
        is_continuation = entry_ts in open_positions  # partial close của lệnh đang mở

        if not is_continuation:
            # Lệnh mới — dọn dẹp lazy trước, rồi kiểm tra overlap

            # Xóa các lệnh đã đóng hoàn toàn VÀ exit_ts đã qua (< entry_ts mới)
            stale = [
                ts for ts, info in open_positions.items()
                if info.get("fully_closed") and info["last_exit_ts"] <= entry_ts
            ]
            for ts in stale:
                del open_positions[ts]

            # ── Rule 2: OPEN_WHILE_OPEN ───────────────────────────────────────
            # Lệnh mới entry_ts < last_exit_ts của lệnh cũ → hai lệnh chồng nhau
            for open_ts, open_info in list(open_positions.items()):
                if entry_ts < open_info["last_exit_ts"]:
                    _add(idx, "OPEN_WHILE_OPEN", "error",
                         f"Lệnh #{idx+1} ({side}) entry={trade.get('entry_time')} "
                         f"overlap với lệnh {open_info['side']} "
                         f"(entry_ts={open_ts}, last_exit={open_info['last_exit_ts']})",
                         trade)
                    break

            # ── Rule 8: DUPLICATE_TIMESTAMP ───────────────────────────────────
            # entry_ts đã thấy trước đó nhưng không còn trong open_positions
            # (lệnh cũ đã đóng hết) → hai lệnh khác nhau cùng entry_ts
            if entry_ts in seen_entry_ts and entry_ts not in open_positions:
                _add(idx, "DUPLICATE_TIMESTAMP", "warning",
                     f"Lệnh #{idx+1}: entry_ts={entry_ts} trùng với lệnh đã đóng trước đó "
                     f"(side={side}, entry={trade.get('entry_time')})",
                     trade)

            open_positions[entry_ts] = {"side": side, "last_exit_ts": exit_ts}
            seen_entry_ts.add(entry_ts)

        else:
            # Partial close hoặc full close của lệnh đang mở
            existing = open_positions[entry_ts]

            # ── Rule 3: SIDE_MISMATCH ─────────────────────────────────────────
            if existing["side"] != side:
                _add(idx, "SIDE_MISMATCH", "error",
                     f"Lệnh #{idx+1}: đóng side='{side}' nhưng lệnh đang mở là "
                     f"side='{existing['side']}' (entry={trade.get('entry_time')})",
                     trade)

            # Cập nhật exit_ts mới nhất để phát hiện overlap với lệnh tiếp theo
            open_positions[entry_ts]["last_exit_ts"] = max(
                open_positions[entry_ts]["last_exit_ts"], exit_ts
            )

        # ── Rule 1: CLOSE_WITHOUT_OPEN ────────────────────────────────────────
        if entry_ts not in open_positions and entry_ts not in seen_entry_ts:
            _add(idx, "CLOSE_WITHOUT_OPEN", "error",
                 f"Lệnh #{idx+1}: đóng lệnh tại {trade.get('exit_time')} "
                 f"nhưng không tìm thấy lệnh mở tương ứng (entry_ts={entry_ts})",
                 trade)

        # Xóa khỏi open_positions khi đóng hoàn toàn.
        # Không xóa ngay — giữ lại để phát hiện overlap với lệnh tiếp theo.
        # Xóa lazy ở đầu vòng lặp tiếp theo (khi entry_ts mới > last_exit_ts cũ).
        # Chỉ xóa ngay nếu còn trade tiếp theo cùng entry_ts (partial close tiếp theo).
        next_same_entry = any(
            t.get("entry_ts_ms") == entry_ts
            for t in trades[idx + 1:]
        )
        # Nếu không còn partial close nào nữa → đánh dấu "fully closed"
        # nhưng GIỮ LẠI trong open_positions để phát hiện overlap với lệnh tiếp theo.
        # Sẽ bị xóa khi lệnh tiếp theo có entry_ts > last_exit_ts (xem đầu vòng lặp).
        if not next_same_entry:
            # Đánh dấu fully_closed để phân biệt với partial
            if entry_ts in open_positions:
                open_positions[entry_ts]["fully_closed"] = True

    # ── UNCLOSED_POSITION: lệnh còn mở khi kết thúc simulation ──────────────
    for ts, info in open_positions.items():
        # Bỏ qua các lệnh đã đóng hoàn toàn (fully_closed=True)
        if info.get("fully_closed"):
            continue
        violations.append({
            "index":          -1,
            "trade_num":      -1,
            "rule":           "UNCLOSED_POSITION",
            "severity":       "warning",
            "message":        (
                f"Lệnh {info['side'].upper()} mở tại ts={ts} "
                f"chưa được đóng khi kết thúc simulation"
            ),
            "trade_snapshot": {"entry_ts_ms": ts, "side": info["side"]},
        })

    # ── Rule 9: BALANCE_DRIFT ─────────────────────────────────────────────────
    if trades:
        sum_pnl       = sum(t.get("pnl", 0.0) for t in trades)
        sum_commission = sum(t.get("commission", 0.0) for t in trades)
        # balance_after của lệnh cuối cùng
        final_balance = trades[-1].get("balance_after", initial_balance)
        # Tính balance kỳ vọng: initial - tổng commission (đã trừ khi entry) + sum_pnl
        # Thực ra: balance_after đã bao gồm tất cả, chỉ cần kiểm tra tính nhất quán
        # Kiểm tra: final_balance ≈ initial_balance + sum(net_pnl) - entry_fees_not_in_pnl
        # Đơn giản hơn: kiểm tra balance_after tăng/giảm đúng theo pnl giữa các lệnh liên tiếp
        drift_errors = []
        for idx in range(1, len(trades)):
            prev_bal = trades[idx - 1].get("balance_after", 0)
            curr_bal = trades[idx].get("balance_after", 0)
            curr_pnl = trades[idx].get("pnl", 0)
            # Với partial close: balance thay đổi đúng bằng net_pnl của lần đóng đó
            # Với full close: tương tự
            # Lưu ý: entry fee đã bị trừ khi mở lệnh (không nằm trong trade record)
            # → chỉ kiểm tra: curr_bal = prev_bal + curr_pnl (nếu cùng lệnh)
            # Nếu khác lệnh (entry_ts khác): có thể có entry fee ở giữa → bỏ qua
            prev_entry = trades[idx - 1].get("entry_ts_ms", 0)
            curr_entry = trades[idx].get("entry_ts_ms", 0)
            if prev_entry == curr_entry:
                # Cùng lệnh (partial close) → balance phải tăng đúng bằng pnl
                expected = round(prev_bal + curr_pnl, 4)
                actual   = round(curr_bal, 4)
                if abs(expected - actual) > 0.01:
                    drift_errors.append((idx, expected, actual, abs(expected - actual)))

        if drift_errors:
            for idx, expected, actual, diff in drift_errors[:3]:
                _add(idx, "BALANCE_DRIFT", "error",
                     f"Lệnh #{idx+1}: balance_after={actual:.4f} nhưng kỳ vọng={expected:.4f} "
                     f"(lệch {diff:.4f} USDT) — có thể lỗi tính toán PnL",
                     trades[idx])

    # ── Tổng hợp ──────────────────────────────────────────────────────────────
    error_count   = sum(1 for v in violations if v["severity"] == "error")
    warning_count = sum(1 for v in violations if v["severity"] == "warning")

    # Balance drift tổng thể
    if trades:
        total_pnl_sum  = sum(t.get("pnl", 0.0) for t in trades)
        final_bal      = trades[-1].get("balance_after", initial_balance)
        # Không thể tính drift chính xác vì entry fee không nằm trong trade record
        # Chỉ ghi nhận để debug
        balance_drift = round(abs(final_bal - (initial_balance + total_pnl_sum)), 4)
    else:
        balance_drift = 0.0

    return {
        "ok":              len(violations) == 0,
        "violation_count": len(violations),
        "violations":      violations,
        "stats": {
            "total_trades":  len(trades),
            "error_count":   error_count,
            "warning_count": warning_count,
            "balance_drift": balance_drift,
        },
    }


async def _run_backtest_engine(bot, exchange, start_ms, end_ms, initial_balance,
                                timeframe_override=None, job_id=None,
                                sl_pct_override=None, tp_pct_override=None):
    """
    Backtest engine với pre-computed indicators (nhanh hơn ~100x).
    Cập nhật progress vào _jobs[job_id] nếu có.
    """
    import pandas as pd

    def _update_progress(pct, msg):
        if job_id and job_id in _jobs:
            _jobs[job_id]["progress"] = pct
            _jobs[job_id]["message"] = msg

    strategy_name = bot.strategy_name
    parameters = dict(bot.parameters or {})
    # Override SL/TP nếu được truyền từ request
    if sl_pct_override is not None:
        parameters["stop_loss_pct"] = float(sl_pct_override)
    if tp_pct_override is not None:
        parameters["take_profit_pct"] = float(tp_pct_override)
    timeframe = timeframe_override or parameters.get("timeframe", "5m")
    leverage = int(parameters.get("leverage", 5))
    position_size_pct = float(parameters.get("position_size_pct", 0.10))
    # Phí giao dịch và trượt giá — đọc từ parameters, fallback về hằng số cũ
    commission_pct = float(parameters.get("commission_pct", COMMISSION))
    slippage_pct   = float(parameters.get("slippage_pct", 0.0))
    lookback = _get_lookback(strategy_name, parameters)

    symbols_raw = bot.symbols or ["BTCUSDT"]
    symbol = _normalize_symbol(symbols_raw[0] if symbols_raw else "BTCUSDT")
    tf_ms = _timeframe_ms(timeframe)

    _update_progress(5, f"Đang kiểm tra dữ liệu {symbol} {timeframe}...")
    logger.info(f"Backtest [{job_id}]: fetch {symbol} {timeframe} {start_ms}→{end_ms}")

    # Warmup: lùi về đúng lookback nến trước start_ms
    warmup_start_ms = start_ms - (lookback * tf_ms)

    # ── Bắt buộc đọc từ DB cache — KHÔNG fallback fetch từ exchange ──────────
    # Lý do: fetch từ exchange chỉ có warmup ngắn → EMA chưa hội tụ → kết quả sai.
    # Mọi trường hợp thiếu data đều raise lỗi rõ ràng để user biết cần refresh cache.
    from src.data.ohlcv_service import get_candles as _db_get_candles, get_data_range as _db_get_range

    db_range = await _db_get_range(strategy_name, symbol, timeframe)

    # ── Kiểm tra 1: DB có data không ─────────────────────────────────────────
    if db_range["count"] == 0 or db_range["min_ts"] is None:
        raise ValueError(
            f"❌ Chưa có dữ liệu cache cho {strategy_name}/{symbol}/{timeframe}.\n"
            f"Vào tab 'Market Data' → nhấn 'Cập nhật tất cả' để tải data trước khi backtest."
        )

    # ── Kiểm tra 2: DB có đủ warmup (min_ts <= warmup_start_ms) ──────────────
    if db_range["min_ts"] > warmup_start_ms:
        warmup_need_date = _to_utc7_str(warmup_start_ms)
        db_min_date      = db_range["min_date"]
        raise ValueError(
            f"❌ Dữ liệu cache không đủ warmup cho {strategy_name}/{symbol}/{timeframe}.\n"
            f"Cần data từ: {warmup_need_date} (lookback={lookback} nến × {timeframe})\n"
            f"DB hiện có từ: {db_min_date}\n"
            f"→ Hãy chạy 'Full Refresh' trong tab Market Data để tải lại 5 năm data."
        )

    # ── Kiểm tra 3: DB có đủ đến end_ms (max_ts >= end_ms - lag cho phép) ────
    # Cho phép lag tối đa 3 nến (để tránh false alarm khi data mới nhất chưa đóng)
    max_lag_ms = tf_ms * 3
    if db_range["max_ts"] < end_ms - max_lag_ms:
        db_max_date   = db_range["max_date"]
        need_end_date = _to_utc7_str(end_ms)
        raise ValueError(
            f"❌ Dữ liệu cache chưa cập nhật đến ngày kết thúc backtest.\n"
            f"Cần data đến: {need_end_date}\n"
            f"DB hiện có đến: {db_max_date}\n"
            f"→ Vào tab 'Market Data' → nhấn 'Cập nhật tất cả' để bổ sung data mới."
        )

    # ── Đọc data từ DB ────────────────────────────────────────────────────────
    _update_progress(8, f"Đọc dữ liệu từ DB cache ({db_range['count']} nến)...")
    all_candles = await _db_get_candles(strategy_name, symbol, timeframe, warmup_start_ms, end_ms)
    logger.info(
        f"Backtest [{job_id}]: đọc {len(all_candles)} nến từ DB "
        f"(cache: {db_range['min_date']} → {db_range['max_date']})"
    )

    # ── Kiểm tra 4: Số nến thực tế vs kỳ vọng (phát hiện gap) ───────────────
    # Số nến kỳ vọng = (end_ms - warmup_start_ms) / tf_ms
    # Nếu thiếu >1% → có gap trong data → kết quả tính toán sẽ sai
    expected_candles = int((end_ms - warmup_start_ms) / tf_ms)
    actual_candles   = len(all_candles)
    if expected_candles > 0:
        gap_pct = (expected_candles - actual_candles) / expected_candles * 100
        if gap_pct > 1.0:
            raise ValueError(
                f"❌ Phát hiện gap trong dữ liệu cache {strategy_name}/{symbol}/{timeframe}.\n"
                f"Kỳ vọng ~{expected_candles} nến, thực tế chỉ có {actual_candles} nến "
                f"(thiếu {gap_pct:.1f}%).\n"
                f"→ Vào tab 'Market Data' → nhấn 'Full Refresh' để tải lại data hoàn chỉnh."
            )

    # Số nến tối thiểu thực sự cần: đủ để tính indicator + 1 nến để bắt đầu simulate
    # Với ADTS: lookback đã tính đúng số ngày D1 cần thiết
    # Với các chiến lược khác: lookback là số nến indicator
    min_required = lookback + 1
    if len(all_candles) < min_required:
        raise ValueError(
            f"Không đủ dữ liệu: có {len(all_candles)} nến, cần ít nhất {min_required}. "
            f"Hãy mở rộng khoảng thời gian backtest hoặc giảm bbwidth_sma_period."
        )

    total_candles = len(all_candles)
    logger.info(f"Backtest [{job_id}]: {total_candles} candles fetched")
    _update_progress(15, f"Đã tải {total_candles} nến. Đang tính indicators...")

    # ── Pre-compute indicators (1 lần cho toàn bộ) ───────────────────────────
    df = pd.DataFrame(all_candles, columns=["timestamp","open","high","low","close","volume"])

    if strategy_name == "sma_macd_cross":
        df = _precompute_sma_macd(df, parameters)
    elif strategy_name in ("sma_macd_cross_v2", "sma_macd_cross_v3", "sma_macd_cross_v4", "sma_macd_cross_v5", "sma_macd_cross_v6"):
        df = _precompute_sma_macd(df, parameters)

        # ── V6: kiểm tra ADX đã hội tụ tại điểm bắt đầu simulate ────────────
        if strategy_name == "sma_macd_cross_v6":
            adx_period = int(parameters.get("adx_period", int(float(os.environ.get("ADX_PERIOD", 14)))))
            adx_warmup_needed = adx_period * 2  # seed point của Wilder smoothing
            # Tìm start_idx (nến đầu tiên >= start_ms) để kiểm tra
            ts_arr_check = df["timestamp"].to_numpy()
            start_idx_check = next(
                (idx for idx, t in enumerate(ts_arr_check) if int(t) >= start_ms),
                len(ts_arr_check)
            )
            if start_idx_check < adx_warmup_needed:
                raise ValueError(
                    f"❌ Không đủ warmup để tính ADX({adx_period}) cho V6.\n"
                    f"Cần ít nhất {adx_warmup_needed} nến warmup trước ngày bắt đầu, "
                    f"nhưng chỉ có {start_idx_check} nến.\n"
                    f"→ Chạy 'Full Refresh' trong tab Market Data để có đủ 5 năm data."
                )
            # Kiểm tra giá trị ADX tại start_idx có hợp lệ không (> 0)
            adx_at_start = float(df["adx"].iloc[start_idx_check])
            if adx_at_start == 0.0:
                raise ValueError(
                    f"❌ ADX({adx_period}) chưa hội tụ tại điểm bắt đầu backtest.\n"
                    f"ADX = 0 tại nến đầu tiên của khoảng backtest (index={start_idx_check}).\n"
                    f"Cần thêm warmup: tăng lookback hoặc chọn ngày bắt đầu muộn hơn.\n"
                    f"→ Chạy 'Full Refresh' trong tab Market Data để có đủ 5 năm data."
                )
    elif strategy_name == "sma_macd_cross_v7":
        # V7 = V1 + BB — precompute SMA+MACD rồi thêm BB
        df = _precompute_sma_macd(df, parameters)
        from src.data.indicators import add_bb_to_df as _add_bb
        df = _add_bb(
            df,
            period=parameters.get("bb_period", 20),
            mult=float(parameters.get("bb_mult", 2.0)),
        )
    elif strategy_name == "adts":
        _update_progress(20, "Đang kiểm tra dữ liệu D1 cho ADTS calibration...")
        # ADTS cần thêm D1 data cho daily calibration
        cfg_adts = ADTSConfig.from_dict(parameters)
        d1_days_needed = cfg_adts.bbwidth_sma_period + cfg_adts.atr_period + 10
        d1_start_ms = warmup_start_ms - (d1_days_needed * 86_400_000)

        # Đọc D1 từ DB cache (strategy_name + "1d")
        d1_db_range = await _db_get_range(strategy_name, symbol, "1d")

        if d1_db_range["count"] == 0 or d1_db_range["min_ts"] is None:
            raise ValueError(
                f"❌ Chưa có dữ liệu D1 cache cho {strategy_name}/{symbol}/1d.\n"
                f"ADTS cần D1 data để tính daily calibration.\n"
                f"→ Vào tab 'Market Data' → nhấn 'Cập nhật tất cả'."
            )
        if d1_db_range["min_ts"] > d1_start_ms:
            raise ValueError(
                f"❌ Dữ liệu D1 cache không đủ cho ADTS calibration.\n"
                f"Cần D1 từ: {_to_utc7_str(d1_start_ms)}\n"
                f"DB D1 có từ: {d1_db_range['min_date']}\n"
                f"→ Chạy 'Full Refresh' trong tab Market Data."
            )
        if d1_db_range["max_ts"] < end_ms - 86_400_000 * 3:
            raise ValueError(
                f"❌ Dữ liệu D1 cache chưa cập nhật đến ngày kết thúc backtest.\n"
                f"DB D1 có đến: {d1_db_range['max_date']}\n"
                f"→ Vào tab 'Market Data' → nhấn 'Cập nhật tất cả'."
            )

        _update_progress(21, f"Đọc D1 từ DB cache ({d1_db_range['count']} ngày)...")
        d1_candles = await _db_get_candles(strategy_name, symbol, "1d", d1_start_ms, end_ms)
        logger.info(
            f"Backtest [{job_id}] ADTS: đọc {len(d1_candles)} D1 candles từ DB "
            f"(need ≥{d1_days_needed})"
        )
        if len(d1_candles) < d1_days_needed:
            raise ValueError(
                f"❌ Không đủ D1 data cho ADTS calibration: có {len(d1_candles)} ngày, "
                f"cần ≥{d1_days_needed} ngày.\n"
                f"→ Chạy 'Full Refresh' trong tab Market Data."
            )
        _update_progress(22, f"Đang tính ADTS indicators + Daily Calibration ({len(d1_candles)} ngày D1)...")
        df = _precompute_adts(df, parameters, d1_candles=d1_candles)
    else:
        # Fallback: dùng strategy.analyze() cho các chiến lược khác
        # (chậm hơn nhưng đúng)
        pass

    _update_progress(25, "Đang simulate giao dịch...")

    # ── Tìm start_idx ─────────────────────────────────────────────────────────
    # Bước 1: tìm nến đầu tiên >= start_ms (ngày người dùng chọn)
    first_date_idx = lookback
    for i, c in enumerate(all_candles):
        if c[0] >= start_ms and i >= lookback:
            first_date_idx = i
            break
    if first_date_idx < lookback:
        first_date_idx = lookback

    # Bước 2: với ADTS, đẩy start_idx về phía sau cho đến khi calib_sideway không NaN
    # Đảm bảo không bắt đầu giao dịch khi calibration chưa sẵn sàng
    if strategy_name == "adts" and "_calib_sideway" in df.columns:
        import numpy as np
        calib_arr = df["_calib_sideway"].to_numpy()
        # Tìm index đầu tiên trong df có calib_sideway hợp lệ (không NaN)
        first_valid_calib_idx = next(
            (idx for idx in range(first_date_idx, len(calib_arr))
             if not (isinstance(calib_arr[idx], float) and np.isnan(calib_arr[idx]))),
            None
        )
        if first_valid_calib_idx is None:
            raise ValueError(
                "Không có nến nào có calib_sideway hợp lệ trong khoảng thời gian backtest. "
                "Hãy mở rộng khoảng thời gian hoặc giảm bbwidth_sma_period."
            )
        start_idx = first_valid_calib_idx
        if start_idx > first_date_idx:
            skipped_candles = start_idx - first_date_idx
            skip_days = skipped_candles * _timeframe_ms(timeframe) / 86400_000
            logger.info(
                f"Backtest [{job_id}] ADTS warmup: bỏ qua {skipped_candles} nến đầu "
                f"(~{skip_days:.1f} ngày) cho đến khi calib_sideway sẵn sàng "
                f"tại idx={start_idx} ts={_to_utc7_str(all_candles[start_idx][0])}"
            )
            _update_progress(26, f"ADTS warmup: bắt đầu simulate từ {_to_utc7_str(all_candles[start_idx][0])}")
    else:
        start_idx = first_date_idx

    # ── Simulation loop ───────────────────────────────────────────────────────
    balance = initial_balance
    open_position = None
    trades = []
    equity_curve = [{"ts": all_candles[start_idx][0], "balance": balance, "pnl_cum": 0.0, "drawdown_pct": 0.0}]
    # Dense equity curve: ghi nhận balance (bao gồm unrealized PnL) tại MỌI nến
    # Dùng để tính MDD chính xác, bao gồm cả drawdown trong lúc đang giữ lệnh
    dense_equity: list[dict] = []
    peak_balance = balance
    last_entry_phase: dict = {}

    simulate_range = [i for i in range(start_idx, len(all_candles)) if all_candles[i][0] <= end_ms]
    total_sim = len(simulate_range)

    for loop_idx, i in enumerate(simulate_range):
        candle = all_candles[i]
        ts_ms = candle[0]

        # Progress update mỗi 5%
        if total_sim > 0 and loop_idx % max(1, total_sim // 20) == 0:
            pct = 25 + int(loop_idx / total_sim * 65)
            _update_progress(pct, f"Simulate nến {loop_idx+1}/{total_sim}...")

        # ── Ghi nhận equity tại mỗi nến (bao gồm unrealized PnL) ─────────────
        # Dùng giá close của nến hiện tại để ước tính unrealized PnL
        candle_close = candle[4]
        if open_position:
            pos_pv = open_position["position_value"]
            pos_ep = open_position["entry_price"]
            if pos_ep > 0:
                if open_position["side"] == "long":
                    unrealized_pct = (candle_close - pos_ep) / pos_ep
                else:
                    unrealized_pct = (pos_ep - candle_close) / pos_ep
                unrealized_pnl = pos_pv * unrealized_pct
            else:
                unrealized_pnl = 0.0
        else:
            unrealized_pnl = 0.0
        equity_now = balance + unrealized_pnl
        dense_equity.append({"ts": ts_ms, "equity": equity_now})

        # ── Lấy signal ────────────────────────────────────────────────────────
        if strategy_name in ("sma_macd_cross", "sma_macd_cross_v2", "sma_macd_cross_v3", "sma_macd_cross_v4", "sma_macd_cross_v5", "sma_macd_cross_v6") and i >= 2:
            sig = _simulate_sma_macd_candle(df, i, open_position, last_entry_phase, parameters, strategy_name)
        elif strategy_name == "sma_macd_cross_v7" and i >= 2:
            sig = _simulate_sma_macd_v7_candle(df, i, open_position, last_entry_phase, parameters)
            sig = _simulate_sma_macd_candle(df, i, open_position, last_entry_phase, parameters, strategy_name)
        elif strategy_name == "adts" and i >= 2:
            sig = _simulate_adts_candle(df, i, open_position, parameters)
        else:
            # Fallback: gọi strategy.analyze() (chậm)
            ohlcv_slice = all_candles[max(0, i - lookback * 2): i + 1]
            sim_pos = []
            if open_position:
                sim_pos = [{"symbol": symbol, "side": open_position["side"],
                            "size": open_position["size"], "entry_price": open_position["entry_price"],
                            "metadata": open_position.get("metadata", {})}]
            try:
                strategy = _build_strategy(strategy_name, parameters)
                signal_obj = await strategy.analyze(symbol, ohlcv_slice, sim_pos)
                sig_type = signal_obj.signal if not signal_obj.is_none else "none"
                sig = {"type": sig_type, "price": signal_obj.price, "metadata": signal_obj.metadata or {}}
            except Exception as e:
                logger.warning(f"Strategy error at candle {i}: {e}")
                continue

        sig_type = sig.get("type", "none")

        # ── Entry ─────────────────────────────────────────────────────────────
        if sig_type in ("long", "short") and open_position is None:
            signal_entry_price = sig.get("price") or candle[4]
            if not signal_entry_price or signal_entry_price <= 0:
                signal_entry_price = candle[4]

            # Áp dụng slippage tại thời điểm khớp lệnh:
            #   Long  → mua cao hơn tín hiệu (bất lợi)
            #   Short → bán thấp hơn tín hiệu (bất lợi)
            if slippage_pct > 0 and sig_type == "long":
                entry_price = signal_entry_price * (1.0 + slippage_pct)
            elif slippage_pct > 0 and sig_type == "short":
                entry_price = signal_entry_price * (1.0 - slippage_pct)
            else:
                entry_price = signal_entry_price

            # V4: dung notional_usdt co dinh thay vi % balance
            notional_usdt = float(parameters.get("notional_usdt", 0))
            if notional_usdt > 0:
                position_value = notional_usdt
            else:
                position_value = balance * position_size_pct * leverage
            size = position_value / entry_price
            # Hoa hồng entry = position_value × commission_pct
            fee = position_value * commission_pct
            # Chi phí trượt giá entry (USDT) = |filled - signal| × size
            entry_slippage_cost = abs(entry_price - signal_entry_price) * size
            open_position = {
                "side": sig_type, "entry_price": entry_price, "size": size,
                "position_value": position_value, "entry_fee": fee,
                "entry_slippage_cost": entry_slippage_cost,
                "entry_ts": ts_ms, "entry_candle_idx": i,
                "metadata": sig.get("metadata") or {},
            }
            # ADTS: lưu SL/TP1/trailing từ signal
            if strategy_name == "adts":
                open_position["stop_loss"]           = sig.get("stop_loss", 0)
                open_position["take_profit_1"]       = sig.get("take_profit_1", 0)
                open_position["trailing_stop"]       = sig.get("trailing_stop_init", 0)
                open_position["atr_at_entry"]        = sig.get("atr_at_entry", 0)
                open_position["tp1_hit"]             = False
                open_position["is_emergency_closed"] = False   # flag chống lặp Emergency Exit
                open_position["amount_remaining_pct"] = 1.0    # 100% còn lại
            balance -= fee
            # One-shot tracking
            phase_ts = (sig.get("metadata") or {}).get("sig_phase_start_ts")
            if phase_ts:
                last_entry_phase[sig_type] = phase_ts

        # ── Exit ──────────────────────────────────────────────────────────────
        elif sig_type in ("close_long", "close_short") and open_position is not None:
            exit_price = sig.get("price") or candle[4]
            if not exit_price or exit_price <= 0:
                exit_price = candle[4]

            # Áp dụng slippage tại thời điểm khớp lệnh:
            #   Đóng Long  → bán thấp hơn tín hiệu (bất lợi)
            #   Đóng Short → mua cao hơn tín hiệu (bất lợi)
            if slippage_pct > 0:
                if sig_type == "close_long":
                    exit_price = exit_price * (1.0 - slippage_pct)
                else:
                    exit_price = exit_price * (1.0 + slippage_pct)

            pos = open_position
            pv = pos["position_value"]
            # Lay leverage thuc te: V4 dung leverage_v4, cac version khac dung leverage
            eff_leverage = float(parameters.get("leverage_v4", 0)) or leverage

            # ADTS: xử lý partial close (TP1 chốt 50%, Emergency chốt 50%)
            is_partial  = sig.get("partial", False)
            partial_pct = float(sig.get("partial_pct", 1.0))

            # ── Kiểm tra min_notional ─────────────────────────────────────────
            # Nếu phần còn lại sau partial close < min_notional → đóng toàn bộ
            # Tránh tình trạng giữ lệnh với notional quá nhỏ (dưới mức tối thiểu sàn)
            if is_partial:
                min_notional = float(parameters.get("min_notional", 5.0))
                remaining_pv = pv * (1.0 - partial_pct)
                if remaining_pv < min_notional:
                    is_partial  = False
                    partial_pct = 1.0
                    # Cập nhật reason để phản ánh việc upgrade lên full close
                    original_reason = sig.get("reason", "")
                    sig = dict(sig)
                    sig["reason"] = (
                        f"{original_reason} → Full Close "
                        f"(còn lại {remaining_pv:.2f} USDT < min_notional {min_notional:.2f} USDT)"
                    )

            # Tính PnL trên phần đóng
            close_pv = pv * partial_pct if is_partial else pv

            if pos["side"] == "long":
                price_change_pct = (exit_price - pos["entry_price"]) / pos["entry_price"]
            else:
                price_change_pct = (pos["entry_price"] - exit_price) / pos["entry_price"]

            gross_pnl = close_pv * price_change_pct
            # Hoa hồng exit = close_pv × commission_pct
            exit_fee  = close_pv * commission_pct
            # entry_fee tính theo tỷ lệ đóng
            entry_fee_portion = pos["entry_fee"] * partial_pct
            net_pnl = gross_pnl - entry_fee_portion - exit_fee

            margin = close_pv / eff_leverage
            pnl_pct = net_pnl / margin * 100 if margin > 0 else 0.0
            balance += net_pnl
            holding_candles = i - pos["entry_candle_idx"]

            trades.append({
                "entry_time": _to_utc7_str(pos["entry_ts"]),
                "exit_time":  _to_utc7_str(ts_ms),
                "entry_ts_ms": pos["entry_ts"],
                "exit_ts_ms":  ts_ms,
                "symbol": symbol, "side": pos["side"],
                "entry_price": pos["entry_price"], "exit_price": exit_price,
                "size": round(pos["size"] * partial_pct, 6),
                "pnl": round(net_pnl, 4), "pnl_pct": round(pnl_pct, 2),
                "balance_after": round(balance, 4),
                "holding_candles": holding_candles,
                "exit_reason": sig.get("reason", ""),
                # Chi phí giao dịch (đã khấu trừ vào net_pnl)
                "commission": round(entry_fee_portion + exit_fee, 6),
                "slippage_cost": round(
                    pos.get("entry_slippage_cost", 0.0) * partial_pct
                    + abs(exit_price - sig.get("price", exit_price)) * pos["size"] * partial_pct,
                    6
                ),
            })
            pnl_cum = balance - initial_balance
            peak_balance = max(peak_balance, balance)
            drawdown_pct = (peak_balance - balance) / peak_balance * 100 if peak_balance > 0 else 0.0
            equity_curve.append({"ts": ts_ms, "balance": round(balance, 4),
                                  "pnl_cum": round(pnl_cum, 4), "drawdown_pct": round(drawdown_pct, 2)})

            if is_partial:
                # Cập nhật position: giảm size và position_value, giữ lệnh
                open_position["position_value"]      = pv * (1.0 - partial_pct)
                open_position["size"]                = pos["size"] * (1.0 - partial_pct)
                open_position["entry_fee"]           = pos["entry_fee"] * (1.0 - partial_pct)
                open_position["entry_slippage_cost"] = pos.get("entry_slippage_cost", 0.0) * (1.0 - partial_pct)
                # Không xóa open_position — lệnh vẫn còn
            else:
                open_position = None

    _update_progress(92, "Đang tính thống kê...")

    # ── Summary stats ─────────────────────────────────────────────────────────
    total_trades = len(trades)
    winning = [t for t in trades if t["pnl"] > 0]
    losing  = [t for t in trades if t["pnl"] <= 0]
    win_count  = len(winning)
    loss_count = len(losing)
    win_rate = round(win_count / total_trades * 100, 2) if total_trades > 0 else 0.0
    total_pnl = sum(t["pnl"] for t in trades)
    total_return_pct = round((balance - initial_balance) / initial_balance * 100, 2)
    gross_profit = sum(t["pnl"] for t in winning)
    gross_loss   = abs(sum(t["pnl"] for t in losing))
    profit_factor = round(gross_profit / gross_loss, 3) if gross_loss > 0 else 999.0
    avg_win  = round(gross_profit / win_count, 4)  if win_count  > 0 else 0.0
    avg_loss = round(-gross_loss / loss_count, 4)  if loss_count > 0 else 0.0
    largest_win  = round(max((t["pnl"] for t in winning), default=0.0), 4)
    largest_loss = round(min((t["pnl"] for t in losing),  default=0.0), 4)
    avg_holding  = round(sum(t["holding_candles"] for t in trades) / total_trades, 1) if total_trades > 0 else 0.0

    # ── MDD từ Equity Curve liên tục (bao gồm unrealized PnL) ────────────────
    # Đây là MDD thực tế mà trader phải chịu đựng khi bot chạy live
    # Khác với MDD từ trade list (chỉ tính tại điểm đóng lệnh)
    if dense_equity:
        eq_values = [e["equity"] for e in dense_equity]
        eq_ts     = [e["ts"]     for e in dense_equity]

        # Tính MDD: duyệt qua từng điểm, theo dõi peak và trough
        peak_eq   = eq_values[0]
        mdd_pct   = 0.0          # Max Drawdown %
        mdd_usdt  = 0.0          # Max Drawdown tuyệt đối (USDT)
        mdd_peak_ts   = eq_ts[0]
        mdd_trough_ts = eq_ts[0]
        mdd_peak_val  = eq_values[0]
        mdd_trough_val = eq_values[0]

        cur_peak     = eq_values[0]
        cur_peak_ts  = eq_ts[0]
        cur_peak_val = eq_values[0]

        for ts_i, eq_i in zip(eq_ts, eq_values):
            if eq_i > cur_peak:
                cur_peak     = eq_i
                cur_peak_ts  = ts_i
                cur_peak_val = eq_i
            dd_pct  = (cur_peak - eq_i) / cur_peak * 100 if cur_peak > 0 else 0.0
            dd_usdt = cur_peak - eq_i
            if dd_pct > mdd_pct:
                mdd_pct        = dd_pct
                mdd_usdt       = dd_usdt
                mdd_peak_ts    = cur_peak_ts
                mdd_trough_ts  = ts_i
                mdd_peak_val   = cur_peak_val
                mdd_trough_val = eq_i

        mdd_pct  = round(mdd_pct, 2)
        mdd_usdt = round(mdd_usdt, 4)

        # MDD Duration: thời gian từ peak → trough (ngày)
        mdd_duration_days = round(
            (mdd_trough_ts - mdd_peak_ts) / 86_400_000, 1
        ) if mdd_trough_ts > mdd_peak_ts else 0.0

        # MDD Recovery: thời gian từ trough đến khi equity vượt lại peak (ngày)
        # Tìm điểm đầu tiên sau trough mà equity >= mdd_peak_val
        recovery_ts = None
        for ts_i, eq_i in zip(eq_ts, eq_values):
            if ts_i > mdd_trough_ts and eq_i >= mdd_peak_val:
                recovery_ts = ts_i
                break
        if recovery_ts is not None:
            mdd_recovery_days = round((recovery_ts - mdd_trough_ts) / 86_400_000, 1)
        else:
            mdd_recovery_days = None  # Chưa hồi phục trong khoảng backtest

        # MDD từ trade list (điểm đóng lệnh) — giữ lại để so sánh
        peak_trade = initial_balance
        max_dd_trade = 0.0
        for t in trades:
            peak_trade = max(peak_trade, t["balance_after"])
            dd = (peak_trade - t["balance_after"]) / peak_trade * 100 if peak_trade > 0 else 0.0
            max_dd_trade = max(max_dd_trade, dd)
        max_drawdown_pct = round(max_dd_trade, 2)  # backward compat

    else:
        # Không có dữ liệu dense equity (không có lệnh nào)
        mdd_pct = 0.0
        mdd_usdt = 0.0
        mdd_duration_days = 0.0
        mdd_recovery_days = None
        mdd_peak_ts = mdd_trough_ts = 0
        mdd_peak_val = mdd_trough_val = initial_balance
        max_drawdown_pct = 0.0
    if len(trades) > 1:
        returns = [t["pnl_pct"] / 100 for t in trades]
        mean_r = sum(returns) / len(returns)
        std_r  = math.sqrt(sum((r - mean_r) ** 2 for r in returns) / len(returns))
        tf_per_day = 86400_000 / tf_ms
        annual_factor = math.sqrt(tf_per_day * 252)
        sharpe = round((mean_r / std_r * annual_factor) if std_r > 0 else 0.0, 3)
    else:
        sharpe = 0.0

    summary = {
        "total_trades": total_trades, "winning_trades": win_count, "losing_trades": loss_count,
        "win_rate": win_rate, "total_pnl": round(total_pnl, 4), "total_return_pct": total_return_pct,
        "max_drawdown_pct": max_drawdown_pct, "profit_factor": profit_factor,
        "avg_win": avg_win, "avg_loss": avg_loss,
        "largest_win": largest_win, "largest_loss": largest_loss,
        "avg_holding_candles": avg_holding, "sharpe_ratio": sharpe,
        "initial_balance": initial_balance, "final_balance": round(balance, 4),
        "total_commission": round(sum(t.get("commission", 0.0) for t in trades), 4),
        "total_slippage_cost": round(sum(t.get("slippage_cost", 0.0) for t in trades), 4),
        "commission_pct": commission_pct,
        "slippage_pct": slippage_pct,
        # MDD từ Equity Curve liên tục
        "mdd_equity_pct":       mdd_pct,
        "mdd_equity_usdt":      mdd_usdt,
        "mdd_duration_days":    mdd_duration_days,
        "mdd_recovery_days":    mdd_recovery_days,
        "mdd_peak_ts":          _to_utc7_str(mdd_peak_ts)   if mdd_peak_ts   else "",
        "mdd_trough_ts":        _to_utc7_str(mdd_trough_ts) if mdd_trough_ts else "",
        "mdd_peak_balance":     round(mdd_peak_val, 4),
        "mdd_trough_balance":   round(mdd_trough_val, 4),
    }
    # ── Integrity validation ──────────────────────────────────────────────────
    integrity = validate_backtest_integrity(trades, initial_balance)
    if integrity["violations"]:
        logger.warning(
            f"Backtest [{job_id}] integrity: {integrity['violation_count']} vi phạm — "
            + "; ".join(v["message"] for v in integrity["violations"][:3])
        )
    summary["integrity"] = integrity

    return {"symbol": symbol, "timeframe": timeframe, "trades": trades,
            "equity_curve": equity_curve, "dense_equity": dense_equity, "summary": summary,
            "chart_data": _build_chart_data(df, all_candles, trades, strategy_name, start_ms, end_ms)}


# ── Chart data builder ───────────────────────────────────────────────────────

def _build_chart_data(df, all_candles: list, trades: list, strategy_name: str,
                      start_ms: int, end_ms: int) -> dict:
    """
    Build chart data từ DataFrame đã có pre-computed indicators.
    Format giống hệt /api/chart-data/{symbol} để dùng lại renderChart logic.
    Chỉ trả về nến trong khoảng [start_ms, end_ms] (bỏ warmup).
    """
    import math

    def _safe(v, decimals=8):
        if v is None:
            return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else round(f, decimals)
        except (TypeError, ValueError):
            return None

    def _str(v, default='yellow'):
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return default
        return str(v)

    candles_out = []

    for idx, row in df.iterrows():
        ts = int(row["timestamp"])
        if ts < start_ms or ts > end_ms:
            continue

        # Format giống hệt market.py /api/chart-data
        candles_out.append({
            "x":    ts,
            "o":    _safe(row["open"],  2),
            "h":    _safe(row["high"],  2),
            "l":    _safe(row["low"],   2),
            "c":    _safe(row["close"], 2),
            # Custom SMA
            "sma_up":           _safe(row.get("custom_sma_up")),
            "sma_dn":           _safe(row.get("custom_sma_dn")),
            "sma_trend":        _safe(row.get("custom_sma_trend"), 0),
            "sma_basis":        _safe(row.get("custom_sma_basis")),
            "sma_momentum":     _str(row.get("custom_sma_momentum"), "yellow"),
            "sma_slope_pct":    _safe(row.get("custom_sma_slope_pct")),
            "sma_momentum_pct": _safe(row.get("custom_sma_momentum_pct")),
            "sma_momentum_n":   _str(row.get("custom_sma_momentum_n"), "yellow"),
            "sma_momentum_n_pct": _safe(row.get("custom_sma_momentum_n_pct")),
            # Custom MACD
            "macd":                 _safe(row.get("custom_macd")),
            "macd_signal":          _safe(row.get("custom_macd_signal")),
            "macd_hist":            _safe(row.get("custom_macd_hist")),
            "macd_hist_color":      _str(row.get("custom_macd_hist_color"), "above_grow"),
            "macd_momentum":        _str(row.get("custom_macd_momentum"), "yellow"),
            "macd_sig_momentum":    _str(row.get("custom_macd_sig_momentum"), "yellow"),
            "macd_slope_pct":       _safe(row.get("custom_macd_slope_pct")),
            "macd_sig_slope_pct":   _safe(row.get("custom_macd_sig_slope_pct")),
            "macd_momentum_pct":    _safe(row.get("custom_macd_momentum_pct")),
            "macd_sig_momentum_pct": _safe(row.get("custom_macd_sig_momentum_pct")),
            # ADX
            "adx":          _safe(row.get("adx"),          4),
            "adx_plus_di":  _safe(row.get("adx_plus_di"),  4),
            "adx_minus_di": _safe(row.get("adx_minus_di"), 4),
        })

    # Trades: entry/exit markers
    trades_out = []
    for t in trades:
        trades_out.append({
            "entry_ts":    t.get("entry_ts_ms"),
            "exit_ts":     t.get("exit_ts_ms"),
            "entry_time":  t.get("entry_time", ""),
            "exit_time":   t.get("exit_time", ""),
            "side":        t.get("side", ""),
            "entry_price": _safe(t.get("entry_price"), 4),
            "exit_price":  _safe(t.get("exit_price"), 4),
            "pnl":         _safe(t.get("pnl"), 4),
            "pnl_pct":     _safe(t.get("pnl_pct"), 2),
            "exit_reason": t.get("exit_reason", ""),
            "holding_candles": t.get("holding_candles", 0),
        })

    return {
        "strategy_name": strategy_name,
        "candles":  candles_out,
        "trades":   trades_out,
        "adx_threshold": float(os.environ.get("ADX_ENTRY_THRESHOLD", 25.0)),
    }


# ── Excel export ─────────────────────────────────────────────────────────────

def _create_excel(bot, result, start_date, end_date, filepath):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Run: pip install openpyxl>=3.1.2")

    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

    wb = openpyxl.Workbook()
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill    = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font   = Font(bold=True)

    summary      = result["summary"]
    trades       = result["trades"]
    equity_curve = result["equity_curve"]
    symbol       = result["symbol"]
    timeframe    = result["timeframe"]

    # ── Sheet 1: Tổng hợp ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tong hop"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 25

    info_rows = [
        ("THONG TIN BOT", ""),
        ("Ten bot",                  bot.name),
        ("Chien luoc",               bot.strategy_name),
        ("Symbol",                   symbol),
        ("Timeframe",                timeframe),
        ("Tu ngay",                  start_date),
        ("Den ngay",                 end_date),
        ("Von ban dau (USDT)",       summary["initial_balance"]),
        ("Hoa hong / lenh (%)",      round(summary.get("commission_pct", 0.0005) * 100, 4)),
        ("Truot gia / lenh (%)",     round(summary.get("slippage_pct", 0.0) * 100, 4)),
        ("", ""),
        ("KET QUA BACKTEST", ""),
        ("Tong so lenh",             summary["total_trades"]),
        ("Lenh thang",               summary["winning_trades"]),
        ("Lenh thua",                summary["losing_trades"]),
        ("Ti le thang (%)",          summary["win_rate"]),
        ("Tong Pnl (USDT)",          summary["total_pnl"]),
        ("Tong loi nhuan (%)",       summary["total_return_pct"]),
        ("Von cuoi (USDT)",          summary["final_balance"]),
        ("Max Drawdown (%)",         summary["max_drawdown_pct"]),
        ("Profit Factor",            summary["profit_factor"]),
        ("TB lenh thang (USDT)",     summary["avg_win"]),
        ("TB lenh thua (USDT)",      summary["avg_loss"]),
        ("Lenh thang lon nhat",      summary["largest_win"]),
        ("Lenh thua lon nhat",       summary["largest_loss"]),
        ("TB thoi gian giu (nen)",   summary["avg_holding_candles"]),
        ("Sharpe Ratio",             summary["sharpe_ratio"]),
        ("", ""),
        ("CHI PHI GIAO DICH", ""),
        ("Tong hoa hong (USDT)",     summary.get("total_commission", 0.0)),
        ("Tong truot gia (USDT)",    summary.get("total_slippage_cost", 0.0)),
        ("Tong chi phi (USDT)",      round(
            summary.get("total_commission", 0.0) + summary.get("total_slippage_cost", 0.0), 4
        )),
        ("", ""),
        ("PHAN TICH MAX DRAWDOWN (EQUITY CURVE)", ""),
        ("MDD Equity Curve (%)",      summary.get("mdd_equity_pct", 0.0)),
        ("MDD Equity Curve (USDT)",   summary.get("mdd_equity_usdt", 0.0)),
        ("MDD tu diem dong lenh (%)", summary.get("max_drawdown_pct", 0.0)),
        ("Thoi gian MDD (ngay)",      summary.get("mdd_duration_days", 0.0)),
        ("Thoi gian hoi phuc (ngay)",
            summary.get("mdd_recovery_days")
            if summary.get("mdd_recovery_days") is not None
            else "Chua hoi phuc trong ky backtest"),
        ("Dinh truoc MDD",            summary.get("mdd_peak_ts", "")),
        ("Day MDD",                   summary.get("mdd_trough_ts", "")),
        ("So du tai dinh (USDT)",     summary.get("mdd_peak_balance", 0.0)),
        ("So du tai day (USDT)",      summary.get("mdd_trough_balance", 0.0)),
    ]

    # Thêm section INTEGRITY nếu có vi phạm
    integrity = summary.get("integrity", {})
    if integrity.get("violation_count", 0) > 0:
        info_rows += [
            ("", ""),
            ("KIEM TRA TINH TOAN VEN (INTEGRITY)", ""),
        ]
        stats = integrity.get("stats", {})
        info_rows += [
            ("Tong vi pham",              integrity.get("violation_count", 0)),
            ("Loi nghiem trong (error)",  stats.get("error_count", 0)),
            ("Canh bao (warning)",        stats.get("warning_count", 0)),
            ("Balance drift (USDT)",      stats.get("balance_drift", 0.0)),
        ]
        for v_idx, v in enumerate(integrity.get("violations", [])[:10], start=1):
            info_rows.append((
                f"  Vi pham #{v_idx} [{v['rule']}]",
                v["message"][:120],
            ))

    pos_labels = {"Tong Pnl (USDT)", "Tong loi nhuan (%)", "Von cuoi (USDT)",
                  "Profit Factor", "Sharpe Ratio", "TB lenh thang (USDT)", "Lenh thang lon nhat",
                  "So du tai dinh (USDT)"}
    neg_labels = {"Max Drawdown (%)", "TB lenh thua (USDT)", "Lenh thua lon nhat",
                  "Tong hoa hong (USDT)", "Tong truot gia (USDT)", "Tong chi phi (USDT)",
                  "MDD Equity Curve (%)", "MDD Equity Curve (USDT)",
                  "MDD tu diem dong lenh (%)", "Thoi gian MDD (ngay)", "Thoi gian hoi phuc (ngay)",
                  "So du tai day (USDT)",
                  "Loi nghiem trong (error)", "Canh bao (warning)", "Balance drift (USDT)"}
    mdd_section_labels = {"PHAN TICH MAX DRAWDOWN (EQUITY CURVE)"}
    integrity_section_labels = {"KIEM TRA TINH TOAN VEN (INTEGRITY)"}

    for row_idx, (label, value) in enumerate(info_rows, start=1):
        cell_a = ws1.cell(row=row_idx, column=1, value=label)
        cell_b = ws1.cell(row=row_idx, column=2, value=value)
        if label in ("THONG TIN BOT", "KET QUA BACKTEST", "CHI PHI GIAO DICH"):
            cell_a.font = header_font; cell_a.fill = header_fill
            cell_b.fill = header_fill
        elif label in mdd_section_labels:
            # MDD section header — màu cam đậm để nổi bật
            mdd_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            cell_a.font = header_font; cell_a.fill = mdd_fill
            cell_b.fill = mdd_fill
        elif label in integrity_section_labels:
            # Integrity section header — màu đỏ nếu có lỗi, xanh lá nếu OK
            has_errors = integrity.get("stats", {}).get("error_count", 0) > 0
            int_color = "C00000" if has_errors else "375623"
            int_fill = PatternFill(start_color=int_color, end_color=int_color, fill_type="solid")
            cell_a.font = header_font; cell_a.fill = int_fill
            cell_b.fill = int_fill
        elif label:
            cell_a.font = bold_font
            if isinstance(value, (int, float)):
                if label in pos_labels:
                    cell_b.fill = green_fill if value >= 0 else red_fill
                elif label in neg_labels:
                    cell_b.fill = red_fill if value != 0 else green_fill

    # ── Sheet 2: Chi tiết lệnh ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Chi tiet lenh")
    headers2   = ["#", "Thoi gian vao", "Thoi gian ra", "Symbol", "Side",
                  "Gia vao", "Gia ra", "So luong", "Pnl (USDT)", "Pnl (%)",
                  "So du sau lenh", "Thoi gian giu (nen)",
                  "Hoa hong (USDT)", "Truot gia (USDT)"]
    col_widths2 = [5, 18, 18, 12, 8, 14, 14, 12, 14, 10, 16, 18, 16, 16]
    for col_idx, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    for t_idx, trade in enumerate(trades, start=1):
        row      = t_idx + 1
        fill     = alt_fill if t_idx % 2 == 0 else None
        pnl_fill = green_fill if trade["pnl"] > 0 else red_fill
        fee_fill = red_fill  # hoa hồng và trượt giá luôn là chi phí (đỏ nhạt)
        values   = [t_idx, trade["entry_time"], trade["exit_time"], trade["symbol"],
                    trade["side"].upper(), trade["entry_price"], trade["exit_price"],
                    round(trade["size"], 4), trade["pnl"], trade["pnl_pct"],
                    trade["balance_after"], trade["holding_candles"],
                    trade.get("commission", 0.0), trade.get("slippage_cost", 0.0)]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            if col_idx in (9, 10):
                cell.fill = pnl_fill
            elif col_idx in (13, 14):
                chosen = fee_fill if (val and val > 0) else fill
                if chosen is not None:
                    cell.fill = chosen
            elif fill:
                cell.fill = fill

    # ── Sheet 3: Đường vốn ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Duong von")
    headers3    = ["Thoi gian", "So du", "Pnl tich luy", "Drawdown (%)"]
    col_widths3 = [18, 16, 16, 14]
    for col_idx, (h, w) in enumerate(zip(headers3, col_widths3), start=1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    for eq_idx, eq in enumerate(equity_curve, start=1):
        row      = eq_idx + 1
        fill     = alt_fill if eq_idx % 2 == 0 else None
        pnl_fill = green_fill if eq["pnl_cum"] >= 0 else red_fill
        dd_fill  = red_fill if eq["drawdown_pct"] > 5 else (alt_fill if eq["drawdown_pct"] > 0 else None)
        for col_idx, (val, f) in enumerate(
            [(_to_utc7_str(eq["ts"]), fill), (eq["balance"], fill),
             (eq["pnl_cum"], pnl_fill), (eq["drawdown_pct"], dd_fill)],
            start=1
        ):
            cell = ws3.cell(row=row, column=col_idx, value=val)
            if f:
                cell.fill = f

    # ── Sheet 4: MDD Equity Curve (dense — mọi nến) ──────────────────────────
    # Ghi nhận equity (bao gồm unrealized PnL) tại từng nến để phân tích MDD
    # Để tránh file quá lớn, downsample: lấy 1 nến mỗi 15 phút (mỗi 3 nến 5m)
    dense_eq = result.get("dense_equity", [])
    if dense_eq:
        mdd_peak_ts_val   = summary.get("mdd_peak_ts", "")
        mdd_trough_ts_val = summary.get("mdd_trough_ts", "")

        ws4 = wb.create_sheet("MDD Equity Curve")
        headers4    = ["Thoi gian", "Equity (USDT)", "Drawdown (%)"]
        col_widths4 = [18, 18, 14]
        for col_idx, (h, w) in enumerate(zip(headers4, col_widths4), start=1):
            cell = ws4.cell(row=1, column=col_idx, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws4.column_dimensions[get_column_letter(col_idx)].width = w

        # Tính drawdown tại từng điểm của dense curve
        peak_eq_val = dense_eq[0]["equity"] if dense_eq else initial_balance
        step = max(1, len(dense_eq) // 10000)  # tối đa 10,000 dòng
        mdd_fill_cell = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for eq_idx, eq in enumerate(dense_eq[::step], start=1):
            eq_val = eq["equity"]
            ts_val = eq["ts"]
            if eq_val > peak_eq_val:
                peak_eq_val = eq_val
            dd_val = round((peak_eq_val - eq_val) / peak_eq_val * 100, 2) if peak_eq_val > 0 else 0.0

            row  = eq_idx + 1
            fill = alt_fill if eq_idx % 2 == 0 else None
            dd_fill_row = (
                mdd_fill_cell if dd_val == summary.get("mdd_equity_pct", -1)
                else red_fill if dd_val > 10
                else (PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid") if dd_val > 5
                      else (alt_fill if dd_val > 0 else None))
            )
            ts_str = _to_utc7_str(ts_val)
            for col_idx, (val, f) in enumerate(
                [(ts_str, fill), (round(eq_val, 4), fill), (dd_val, dd_fill_row)],
                start=1
            ):
                cell = ws4.cell(row=row, column=col_idx, value=val)
                if f:
                    cell.fill = f

        # Đánh dấu dòng MDD peak và trough bằng comment
        ws4.cell(row=1, column=4, value="Ghi chu")
        ws4.column_dimensions["D"].width = 30
        ws4.cell(row=1, column=4).font = header_font
        ws4.cell(row=1, column=4).fill = header_fill

    wb.save(filepath)
    logger.info(f"Excel saved: {filepath}")


# ── Background job runner ─────────────────────────────────────────────────────

async def _run_job(job_id: str, bot, account, req):
    """Chạy backtest trong background, cập nhật _jobs[job_id]."""
    parameters = bot.parameters or {}
    market_type = parameters.get("market_type", "futures")

    if account:
        exchange = BinanceExchange(
            api_key=account.api_key, api_secret=account.api_secret,
            mode=account.mode, market_type=market_type,
        )
    else:
        exchange = create_exchange_from_env()
        exchange.market_type = market_type

    try:
        start_dt = datetime.strptime(req.start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        start_ms = int(start_dt.timestamp() * 1000)
        if req.end_date:
            end_dt = datetime.strptime(req.end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
        else:
            end_dt = datetime.now(timezone.utc)
        end_ms = int(end_dt.timestamp() * 1000)
        end_date_str = end_dt.strftime("%Y-%m-%d")

        # Validate: start phải trước end
        if start_ms >= end_ms:
            raise ValueError(
                f"Ngày bắt đầu ({req.start_date}) phải trước ngày kết thúc ({end_date_str}). "
                f"Kiểm tra lại khoảng thời gian."
            )

        # Retry connect toi da 3 lan neu timeout
        connect_ok = False
        last_connect_err = None
        for attempt in range(1, 4):
            try:
                _jobs[job_id]["message"] = f"Đang kết nối exchange (lần {attempt}/3)..."
                await exchange.connect()
                connect_ok = True
                break
            except Exception as e_conn:
                last_connect_err = e_conn
                logger.warning(f"Backtest [{job_id}]: connect attempt {attempt} failed: {e_conn}")
                if attempt < 3:
                    await _asyncio.sleep(3)
                    # Tao lai exchange object de reset connection state
                    if account:
                        exchange = BinanceExchange(
                            api_key=account.api_key, api_secret=account.api_secret,
                            mode=account.mode, market_type=market_type,
                        )
                    else:
                        exchange = create_exchange_from_env()
                        exchange.market_type = market_type

        if not connect_ok:
            raise ValueError(
                f"Không thể kết nối exchange sau 3 lần thử. "
                f"Lỗi: {type(last_connect_err).__name__}: {last_connect_err}. "
                f"Vui lòng thử lại sau."
            )

        result = await _run_backtest_engine(
            bot=bot, exchange=exchange,
            start_ms=start_ms, end_ms=end_ms,
            initial_balance=req.initial_balance,
            timeframe_override=req.timeframe or None,
            sl_pct_override=req.stop_loss_pct,
            tp_pct_override=req.take_profit_pct,
            job_id=job_id,
        )
        await exchange.close()

        _jobs[job_id]["message"] = "Đang xuất Excel..."
        _jobs[job_id]["progress"] = 95

        symbol_safe = result["symbol"].replace("/", "")
        tf_safe     = result["timeframe"]
        start_safe  = req.start_date.replace("-", "")
        end_safe    = end_date_str.replace("-", "")
        filename = f"backtest_{req.bot_id}_{symbol_safe}_{tf_safe}_{start_safe}_{end_safe}.xlsx"
        filepath = os.path.join(BACKTEST_DIR, filename)
        _create_excel(bot=bot, result=result, start_date=req.start_date,
                      end_date=end_date_str, filepath=filepath)

        _jobs[job_id].update({
            "status": "done",
            "progress": 100,
            "message": f"Hoàn tất — {result['summary']['total_trades']} lệnh",
            "result": {
                "success": True,
                "bot_id": req.bot_id, "bot_name": bot.name,
                "symbol": result["symbol"], "timeframe": result["timeframe"],
                "start_date": req.start_date, "end_date": end_date_str,
                "initial_balance": req.initial_balance,
                "summary": result["summary"],
                "trades": result["trades"],
                "equity_curve": result["equity_curve"],
                "excel_filename": filename,
                "download_url": f"/api/backtest/download/{filename}",
            },
            "chart_data": result.get("chart_data"),
        })
        logger.info(f"Backtest job {job_id} done: {result['summary']['total_trades']} trades")

    except Exception as e:
        try:
            await exchange.close()
        except Exception:
            pass
        logger.exception(f"Backtest job {job_id} error: {e}")
        _jobs[job_id].update({
            "status": "error",
            "progress": 0,
            "message": f"Lỗi: {type(e).__name__}: {str(e)}",
            "error": str(e),
        })


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/run")
async def run_backtest(req: BacktestRequest):
    """Khởi động backtest dưới dạng background job. Trả về job_id để poll progress."""

    # Validate ngày trước khi tạo job
    try:
        start_dt = datetime.strptime(req.start_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail=f"start_date không hợp lệ: {req.start_date}")

    if req.end_date:
        try:
            end_dt = datetime.strptime(req.end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"end_date không hợp lệ: {req.end_date}")
        if start_dt >= end_dt:
            raise HTTPException(
                status_code=400,
                detail=f"Ngày bắt đầu ({req.start_date}) phải trước ngày kết thúc ({req.end_date})."
            )

    if req.initial_balance <= 0:
        raise HTTPException(status_code=400, detail="Vốn ban đầu phải > 0")

    async with get_db() as db:
        result = await db.execute(select(Bot).where(Bot.id == req.bot_id, Bot.is_deleted == False))
        bot = result.scalar_one_or_none()
        if not bot:
            raise HTTPException(status_code=404, detail=f"Bot ID={req.bot_id} not found")
        account = None
        if bot.account_id:
            acc_result = await db.execute(select(ExchangeAccount).where(ExchangeAccount.id == bot.account_id))
            account = acc_result.scalar_one_or_none()

    job_id = str(uuid.uuid4())[:8]
    _jobs[job_id] = {
        "status": "running",
        "progress": 0,
        "message": "Đang khởi động...",
        "result": None,
        "error": None,
    }

    # Chạy background — không block HTTP response
    _asyncio.create_task(_run_job(job_id, bot, account, req))

    return {"job_id": job_id, "status": "running"}


@router.get("/progress/{job_id}")
async def get_progress(job_id: str):
    """Poll tiến độ backtest. UI gọi mỗi 2s."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
    return {
        "job_id": job_id,
        "status": job["status"],       # "running" | "done" | "error"
        "progress": job["progress"],   # 0-100
        "message": job["message"],
        "result": job["result"],       # None khi đang chạy, dict khi done
        "error": job["error"],
    }


@router.post("/run-strategy")
async def run_strategy_backtest(req: StrategyBacktestRequest):
    """
    Chạy backtest theo chiến lược + cặp tiền.
    Không cần bot_id — dùng account của bot đầu tiên có sẵn.
    """
    # Validate ngày
    try:
        start_dt = datetime.strptime(req.start_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail=f"start_date không hợp lệ: {req.start_date}")
    if req.end_date:
        try:
            end_dt_check = datetime.strptime(req.end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail=f"end_date không hợp lệ: {req.end_date}")
        if start_dt >= end_dt_check:
            raise HTTPException(status_code=400, detail="start_date phải trước end_date")

    # Lấy account từ bot đầu tiên có sẵn
    async with get_db() as db:
        result = await db.execute(
            select(Bot).where(Bot.is_deleted == False, Bot.status == "running").limit(1)
        )
        ref_bot = result.scalar_one_or_none()
        if not ref_bot:
            result2 = await db.execute(select(Bot).where(Bot.is_deleted == False).limit(1))
            ref_bot = result2.scalar_one_or_none()

        account = None
        if ref_bot and ref_bot.account_id:
            acc_result = await db.execute(
                select(ExchangeAccount).where(ExchangeAccount.id == ref_bot.account_id)
            )
            account = acc_result.scalar_one_or_none()

    # Build params từ request — dùng mặc định của chiến lược nếu không điền
    STRATEGY_DEFAULTS = {
        "sma_macd_cross":    {"bb_length": 200, "timeframe": "5m", "leverage": 5, "position_size_pct": 0.1, "commission_pct": 0.0005, "slippage_pct": 0.0},
        "sma_macd_cross_v2": {"bb_length": 150, "timeframe": "5m", "leverage": 5, "position_size_pct": 0.1, "use_trend_filter": True, "commission_pct": 0.0005, "slippage_pct": 0.0},
        "sma_macd_cross_v3": {"bb_length": 200, "timeframe": "5m", "leverage": 5, "position_size_pct": 0.1, "use_trend_filter": True, "min_ma_distance_pct": 0.1, "min_hold_candles": 3, "commission_pct": 0.0005, "slippage_pct": 0.0},
        "sma_macd_cross_v4": {"bb_length": 200, "timeframe": "5m", "leverage_v4": 10, "notional_usdt": 2000.0, "stop_loss_pct": 3.0, "take_profit_pct": 3.0, "commission_pct": 0.0005, "slippage_pct": 0.0},
        "sma_macd_cross_v5": {"bb_length": 200, "timeframe": "5m", "leverage_v4": 10, "notional_usdt": 2000.0, "stop_loss_pct": 3.0, "take_profit_pct": 3.0, "commission_pct": 0.0005, "slippage_pct": 0.0},
        "sma_macd_cross_v6": {"bb_length": 200, "timeframe": "5m", "leverage": 5, "position_size_pct": 0.1, "adx_period": int(float(os.environ.get("ADX_PERIOD", 14))), "commission_pct": 0.0005, "slippage_pct": 0.0},
        "sma_macd_cross_v7": {"bb_length": 200, "timeframe": "5m", "leverage": 5, "position_size_pct": 0.1, "bb_period": 20, "bb_mult": 2.0, "commission_pct": 0.0005, "slippage_pct": 0.0},
        "adts": {
            "timeframe": "5m", "leverage": 5, "position_size_pct": 0.1,
            "atr_period": 14, "adx_period": 14, "ema_period": 20,
            "ema200_period": 200,
            "bb_period": 20, "bb_std": 2.0, "bbwidth_sma_period": 200,
            "adx_threshold": 20.0, "bbwidth_threshold_factor": 1.0,
            "min_slope_atr_factor": 0.05, "risk_pct": 0.01,
            "sl_atr_mult": 1.5, "hard_sl_pct": 0.03,
            "tp1_rr": 1.2, "tp1_close_pct": 0.5,
            "tp2_trail_atr_mult": 2.0, "emergency_adx_threshold": 20.0,
            "emergency_close_pct": 0.5,
            "min_notional": 5.0,
            "commission_pct": 0.0005, "slippage_pct": 0.0,
        },
    }
    if req.strategy_name not in STRATEGY_DEFAULTS:
        raise HTTPException(status_code=400, detail=f"Chiến lược không hỗ trợ: {req.strategy_name}. Chọn: {list(STRATEGY_DEFAULTS.keys())}")

    params = dict(STRATEGY_DEFAULTS[req.strategy_name])
    params.update({
        "fast_len": 1, "slow_len": 5, "len_c": 200, "factor": 0.05,
        "macd_fast": 12, "macd_slow": 26, "macd_signal_length": 500,
        "macd_src": "EMA", "macd_sig_type": "EMA",
        "lookback_candles": 600, "market_type": "futures",
    })
    # Override từ request nếu có
    if req.timeframe:       params["timeframe"] = req.timeframe
    if req.bb_length:       params["bb_length"] = req.bb_length
    if req.use_trend_filter is not None: params["use_trend_filter"] = req.use_trend_filter
    if req.min_ma_distance_pct is not None: params["min_ma_distance_pct"] = req.min_ma_distance_pct
    if req.min_hold_candles is not None: params["min_hold_candles"] = req.min_hold_candles
    if req.leverage_v4:     params["leverage_v4"] = req.leverage_v4
    if req.notional_usdt:   params["notional_usdt"] = req.notional_usdt
    if req.stop_loss_pct is not None:   params["stop_loss_pct"] = req.stop_loss_pct
    if req.take_profit_pct is not None: params["take_profit_pct"] = req.take_profit_pct
    # ADTS params
    if req.adts_atr_period is not None:               params["atr_period"] = req.adts_atr_period
    if req.adts_adx_period is not None:               params["adx_period"] = req.adts_adx_period
    if req.adts_ema_period is not None:               params["ema_period"] = req.adts_ema_period
    if req.adts_ema200_period is not None:            params["ema200_period"] = req.adts_ema200_period
    if req.adts_bb_period is not None:                params["bb_period"] = req.adts_bb_period
    if req.adts_bbwidth_sma_period is not None:       params["bbwidth_sma_period"] = req.adts_bbwidth_sma_period
    if req.adts_adx_threshold is not None:            params["adx_threshold"] = req.adts_adx_threshold
    if req.adts_bbwidth_threshold_factor is not None: params["bbwidth_threshold_factor"] = req.adts_bbwidth_threshold_factor
    if req.adts_risk_pct is not None:                 params["risk_pct"] = req.adts_risk_pct
    if req.adts_sl_atr_mult is not None:              params["sl_atr_mult"] = req.adts_sl_atr_mult
    if req.adts_hard_sl_pct is not None:              params["hard_sl_pct"] = req.adts_hard_sl_pct
    if req.adts_tp1_rr is not None:                   params["tp1_rr"] = req.adts_tp1_rr
    if req.adts_tp2_trail_atr_mult is not None:       params["tp2_trail_atr_mult"] = req.adts_tp2_trail_atr_mult
    if req.adts_emergency_adx_threshold is not None:  params["emergency_adx_threshold"] = req.adts_emergency_adx_threshold
    if req.adts_leverage is not None:                 params["leverage"] = req.adts_leverage
    if req.adts_min_notional is not None:             params["min_notional"] = req.adts_min_notional
    # V6 params
    if req.adx_entry_threshold is not None: params["adx_entry_threshold"] = req.adx_entry_threshold
    if req.adx_exit_threshold  is not None: params["adx_exit_threshold"]  = req.adx_exit_threshold
    # V7 params
    if req.bb_period is not None: params["bb_period"] = req.bb_period
    if req.bb_mult   is not None: params["bb_mult"]   = req.bb_mult
    # Phí giao dịch & trượt giá (áp dụng cho mọi chiến lược)
    if req.commission_pct is not None: params["commission_pct"] = req.commission_pct
    if req.slippage_pct is not None:   params["slippage_pct"]   = req.slippage_pct
    class _FakeBot:
        strategy_name = req.strategy_name
        parameters = params
        symbols = [req.symbol]
        name = f"{req.strategy_name} / {req.symbol}"
        id = 0

    fake_bot = _FakeBot()

    # Tạo job và chạy
    job_id = str(uuid.uuid4())[:8]
    _jobs[job_id] = {"status": "running", "progress": 0, "message": "Đang khởi động...", "result": None, "error": None}

    _asyncio.create_task(_run_job_strategy(job_id, fake_bot, account, req))
    return {"job_id": job_id, "status": "running"}


async def _run_job_strategy(job_id: str, fake_bot, account, req: StrategyBacktestRequest):
    """Background job cho run-strategy."""
    parameters = fake_bot.parameters
    market_type = parameters.get("market_type", "futures")

    if account:
        exchange = BinanceExchange(
            api_key=account.api_key, api_secret=account.api_secret,
            mode=account.mode, market_type=market_type,
        )
    else:
        exchange = create_exchange_from_env()
        exchange.market_type = market_type

    try:
        start_dt = datetime.strptime(req.start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        start_ms = int(start_dt.timestamp() * 1000)
        if req.end_date:
            end_dt = datetime.strptime(req.end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
        else:
            end_dt = datetime.now(timezone.utc)
        end_ms = int(end_dt.timestamp() * 1000)
        end_date_str = end_dt.strftime("%Y-%m-%d")

        # Retry connect
        connect_ok = False
        last_err = None
        for attempt in range(1, 4):
            try:
                _jobs[job_id]["message"] = f"Đang kết nối exchange (lần {attempt}/3)..."
                await exchange.connect()
                connect_ok = True
                break
            except Exception as e:
                last_err = e
                if attempt < 3:
                    await _asyncio.sleep(3)
                    if account:
                        exchange = BinanceExchange(account.api_key, account.api_secret, account.mode, market_type)
                    else:
                        exchange = create_exchange_from_env()
                        exchange.market_type = market_type
        if not connect_ok:
            raise ValueError(f"Không thể kết nối exchange: {last_err}")

        result = await _run_backtest_engine(
            bot=fake_bot, exchange=exchange,
            start_ms=start_ms, end_ms=end_ms,
            initial_balance=req.initial_balance,
            job_id=job_id,
        )
        await exchange.close()

        _jobs[job_id]["message"] = "Đang xuất Excel..."
        _jobs[job_id]["progress"] = 95

        symbol_safe = result["symbol"].replace("/", "")
        tf_safe = result["timeframe"]
        start_safe = req.start_date.replace("-", "")
        end_safe = end_date_str.replace("-", "")
        filename = f"backtest_{req.strategy_name}_{symbol_safe}_{tf_safe}_{start_safe}_{end_safe}.xlsx"
        filepath = os.path.join(BACKTEST_DIR, filename)
        _create_excel(bot=fake_bot, result=result, start_date=req.start_date, end_date=end_date_str, filepath=filepath)

        _jobs[job_id].update({
            "status": "done", "progress": 100,
            "message": f"Hoàn tất — {result['summary']['total_trades']} lệnh",
            "result": {
                "success": True,
                "strategy_name": req.strategy_name,
                "symbol": result["symbol"], "timeframe": result["timeframe"],
                "start_date": req.start_date, "end_date": end_date_str,
                "initial_balance": req.initial_balance,
                "summary": result["summary"],
                "trades": result["trades"],
                "equity_curve": result["equity_curve"],
                "excel_filename": filename,
                "download_url": f"/api/backtest/download/{filename}",
            },
            "chart_data": result.get("chart_data"),
        })
    except Exception as e:
        try:
            await exchange.close()
        except Exception:
            pass
        logger.exception(f"Strategy backtest job {job_id} error: {e}")
        _jobs[job_id].update({"status": "error", "progress": 0, "message": f"Lỗi: {type(e).__name__}: {str(e)}", "error": str(e)})


@router.get("/chart-data/{job_id}")
async def get_chart_data(job_id: str):
    """Trả về chart data (candles + indicators + trades) cho một backtest job."""
    job = _jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' không tồn tại")
    if job.get("status") != "done":
        raise HTTPException(status_code=400, detail=f"Job chưa hoàn tất (status={job.get('status')})")
    chart_data = job.get("chart_data")
    if not chart_data:
        raise HTTPException(status_code=404, detail="Không có chart data cho job này")
    return chart_data


@router.get("/config")
async def get_backtest_config():
    """Trả về config cho frontend (BACKTEST_CHART_CANDLES, v.v.)."""
    import os
    return {
        "chart_candles": int(os.getenv("BACKTEST_CHART_CANDLES", "200")),
    }


@router.get("/download/{filename}")
async def download_backtest(filename: str):
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    filepath = os.path.join(BACKTEST_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
    return FileResponse(
        path=filepath, filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


