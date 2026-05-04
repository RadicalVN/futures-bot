import openpyxl
f = 'data/backtest/backtest_7_BTCUSDT_5m_20260420_20260430.xlsx'
wb = openpyxl.load_workbook(f)
ws1 = wb['Tong hop']
rows = [(r[0], r[1]) for r in ws1.iter_rows(values_only=True) if r[0]]
for k, v in rows:
    print(f'{k}: {v}')
print('---')
ws2 = wb['Chi tiet lenh']
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[0]:
        print(f"#{row[0]} {row[1]}->{row[2]} {row[4]} entry={row[5]:.2f} exit={row[6]:.2f} pnl={row[8]:.4f}({row[9]:+.2f}%) hold={row[11]}")
