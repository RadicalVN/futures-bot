"""
adts_grid_search.py — Grid Search tối ưu tham số ADTS cho BTC/USDT

Tìm tổ hợp tối ưu nhất dựa trên Sharpe Ratio, Profit Factor, Win Rate, MDD.

Tối ưu hóa:
  - Pre-compute indicators 1 lần, reuse cho tất cả 192 combinations
  - Build ADTSConfig 1 lần mỗi combination (không rebuild mỗi nến)
  - Dùng Binance MAINNET để fetch dữ liệu lịch sử thực

Grid: sl_atr_mult × tp1_rr × adx_threshold × bbwidth_threshold_factor

Chạy:
    python scripts/adts_grid_search.py

Kết quả: data/backtest/adts_grid_search_BTCUSDT_<date>.xlsx
"""
from __future__ import annotations

import asyncio
import itertools
import math
import os
import sys
from datetime import datetime, timezone, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv()

from loguru import logger
from src.dashboard.routers.backtest import (
    _fetch_ohlcv_range,
    _precompute_adts,
    _simulate_adts_candle,
    _timeframe_ms,
)
from src.strategies.adts.models import ADTSConfig

# ── Grid parameters ───────────────────────────────────────────────────────────

SYMBOL          = "BTC/USDT"
TIMEFRAME       = "5m"
START_DATE      = "2026-01-01"
END_DATE        = "2026-04-30"
INITIAL_BALANCE = 10_000.0

SL_ATR_MULT_GRID    = [1.0, 1.5, 2.0, 2.5]
TP1_RR_GRID         = [1.0, 1.2, 1.5, 2.0]
ADX_THRESHOLD_GRID  = [18.0, 20.0, 22.0, 25.0]
BBWIDTH_FACTOR_GRID = [0.85, 1.0, 1.1]

# Tham số cố định — không ảnh hưởng đến pre-compute indicators
FIXED_PARAMS = {
    "timeframe":               TIMEFRAME,
    "leverage":                5,
    "position_size_pct":       0.1,
    "atr_period":              14,
    "adx_period":              14,
    "ema_period":              20,
    "ema200_period":           200,
    "bb_period":               20,
    "bb_std":                  2.0,
    "bbwidth_sma_period":      200,
    "min_slope_atr_factor":    0.05,
    "risk_pct":                0.01,
    "hard_sl_pct":             0.03,
    "tp1_close_pct":           0.5,
    "tp2_trail_atr_mult":      2.0,
    "emergency_adx_threshold": 20.0,
    "emergency_close_pct":     0.5,
    "min_notional":            5.0,
    "commission_pct":          0.0005,
    "slippage_pct":            0.0,
}


# ── Single simulation run ─────────────────────────────────────────────────────

def _run_one(df_computed, all_candles, start_ms, end_ms, params):
    """
    Chạy simulation với df đã pre-computed.
    cfg được build 1 lần, không rebuild mỗi nến.
    """
    import numpy as np

    cfg               = ADTSConfig.from_dict(params)
    commission_pct    = float(params.get("commission_pct", 0.0005))
    slippage_pct      = float(params.get("slippage_pct", 0.0))
    leverage          = int(params.get("leverage", 5))
    position_size_pct = float(params.get("position_size_pct", 0.1))
    tf_ms             = _timeframe_ms(TIMEFRAME)

    # Tìm start_idx — đảm bảo calib_sideway sẵn sàng
    calib_arr   = df_computed["_calib_sideway"].to_numpy()
    first_valid = None
    for idx in range(len(all_candles)):
        if all_candles[idx][0] >= start_ms and idx >= 2:
            for j in range(idx, len(calib_arr)):
                v = calib_arr[j]
                if not (isinstance(v, float) and np.isnan(v)):
                    first_valid = j
                    break
            break
    if first_valid is None:
        return None

    balance       = INITIAL_BALANCE
    open_position = None
    trades        = []
    dense_equity  = []

    simulate_range = [i for i in range(first_valid, len(all_candles))
                      if all_candles[i][0] <= end_ms]

    for i in simulate_range:
        candle = all_candles[i]

        # Dense equity (unrealized PnL)
        if open_position:
            ep = open_position.get("entry_price", 0)
            pv = open_position.get("position_value", 0)
            if ep > 0:
                side = open_position["side"]
                pct  = (candle[4] - ep) / ep if side == "long" else (ep - candle[4]) / ep
                unrealized = pv * pct
            else:
                unrealized = 0.0
        else:
            unrealized = 0.0
        dense_equity.append(balance + unrealized)

        if i < 2:
            continue

        # Truyền cfg pre-built để tránh rebuild Pydantic mỗi nến
        sig      = _simulate_adts_candle(df_computed, i, open_position, params, _cfg=cfg)
        sig_type = sig.get("type", "none")

        # Entry
        if sig_type in ("long", "short") and open_position is None:
            ep = sig.get("price") or candle[4]
            if slippage_pct > 0:
                ep = ep * (1 + slippage_pct) if sig_type == "long" else ep * (1 - slippage_pct)
            pv   = balance * position_size_pct * leverage
            size = pv / ep
            fee  = pv * commission_pct
            open_position = {
                "side": sig_type, "entry_price": ep, "size": size,
                "position_value": pv, "entry_fee": fee,
                "entry_slippage_cost": abs(ep - (sig.get("price") or candle[4])) * size,
                "entry_ts": candle[0], "entry_candle_idx": i,
                "stop_loss":      sig.get("stop_loss", 0),
                "take_profit_1":  sig.get("take_profit_1", 0),
                "trailing_stop":  sig.get("trailing_stop_init", 0),
                "atr_at_entry":   sig.get("atr_at_entry", 0),
                "tp1_hit":        False,
                "is_emergency_closed": False,
                "metadata":       sig.get("metadata") or {},
            }
            balance -= fee

        # Exit
        elif sig_type in ("close_long", "close_short") and open_position is not None:
            xp = sig.get("price") or candle[4]
            if slippage_pct > 0:
                xp = xp * (1 - slippage_pct) if sig_type == "close_long" else xp * (1 + slippage_pct)

            pos         = open_position
            pv          = pos["position_value"]
            is_partial  = sig.get("partial", False)
            partial_pct = float(sig.get("partial_pct", 1.0))

            if is_partial:
                min_notional = float(params.get("min_notional", 5.0))
                if pv * (1.0 - partial_pct) < min_notional:
                    is_partial  = False
                    partial_pct = 1.0

            close_pv  = pv * partial_pct if is_partial else pv
            price_chg = ((xp - pos["entry_price"]) / pos["entry_price"]
                         if pos["side"] == "long"
                         else (pos["entry_price"] - xp) / pos["entry_price"])

            gross_pnl = close_pv * price_chg
            exit_fee  = close_pv * commission_pct
            net_pnl   = gross_pnl - pos["entry_fee"] * partial_pct - exit_fee
            margin    = close_pv / leverage
            pnl_pct   = net_pnl / margin * 100 if margin > 0 else 0.0
            balance  += net_pnl

            trades.append({
                "pnl":             round(net_pnl, 4),
                "pnl_pct":         round(pnl_pct, 2),
                "balance_after":   round(balance, 4),
                "holding_candles": i - pos["entry_candle_idx"],
            })

            if is_partial:
                open_position["position_value"]      = pv * (1.0 - partial_pct)
                open_position["size"]                = pos["size"] * (1.0 - partial_pct)
                open_position["entry_fee"]           = pos["entry_fee"] * (1.0 - partial_pct)
                open_position["entry_slippage_cost"] = pos.get("entry_slippage_cost", 0.0) * (1.0 - partial_pct)
            else:
                open_position = None

    if not trades:
        return None

    total         = len(trades)
    winning       = [t for t in trades if t["pnl"] > 0]
    losing        = [t for t in trades if t["pnl"] <= 0]
    win_rate      = round(len(winning) / total * 100, 2)
    total_pnl     = sum(t["pnl"] for t in trades)
    total_return  = round((balance - INITIAL_BALANCE) / INITIAL_BALANCE * 100, 2)
    gross_profit  = sum(t["pnl"] for t in winning)
    gross_loss    = abs(sum(t["pnl"] for t in losing))
    profit_factor = round(gross_profit / gross_loss, 3) if gross_loss > 0 else 999.0
    avg_holding   = round(sum(t["holding_candles"] for t in trades) / total, 1)

    # MDD từ dense equity
    peak_eq = dense_equity[0] if dense_equity else INITIAL_BALANCE
    mdd_pct = 0.0
    for eq in dense_equity:
        if eq > peak_eq:
            peak_eq = eq
        dd = (peak_eq - eq) / peak_eq * 100 if peak_eq > 0 else 0.0
        mdd_pct = max(mdd_pct, dd)

    # Sharpe
    if len(trades) > 1:
        returns  = [t["pnl_pct"] / 100 for t in trades]
        mean_r   = sum(returns) / len(returns)
        std_r    = math.sqrt(sum((r - mean_r) ** 2 for r in returns) / len(returns))
        tf_per_day = 86_400_000 / tf_ms
        annual_f   = math.sqrt(tf_per_day * 252)
        sharpe = round((mean_r / std_r * annual_f) if std_r > 0 else 0.0, 3)
    else:
        sharpe = 0.0

    return {
        "total_trades":        total,
        "win_rate":            win_rate,
        "total_pnl":           round(total_pnl, 4),
        "total_return_pct":    total_return,
        "profit_factor":       profit_factor,
        "sharpe_ratio":        sharpe,
        "max_drawdown_pct":    round(mdd_pct, 2),
        "avg_holding_candles": avg_holding,
        "final_balance":       round(balance, 4),
    }


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    import numpy as np
    import pandas as pd
    import ccxt as _ccxt

    total_combos = (len(SL_ATR_MULT_GRID) * len(TP1_RR_GRID) *
                    len(ADX_THRESHOLD_GRID) * len(BBWIDTH_FACTOR_GRID))

    logger.info("=" * 70)
    logger.info("ADTS Grid Search — BTC/USDT")
    logger.info(f"sl_atr_mult:          {SL_ATR_MULT_GRID}")
    logger.info(f"tp1_rr:               {TP1_RR_GRID}")
    logger.info(f"adx_threshold:        {ADX_THRESHOLD_GRID}")
    logger.info(f"bbwidth_factor:       {BBWIDTH_FACTOR_GRID}")
    logger.info(f"Period:               {START_DATE} → {END_DATE}")
    logger.info(f"Total combinations:   {total_combos}")
    logger.info("=" * 70)

    # ── Fetch data từ Binance MAINNET (async ccxt, chỉ đọc) ──────────────────
    from src.core.exchange import BinanceExchange
    # Dùng mainnet để lấy dữ liệu lịch sử thực
    # API key không cần thiết cho public endpoints (fetch_ohlcv)
    exchange = BinanceExchange(
        api_key="",
        api_secret="",
        mode="mainnet",
        market_type="futures",
    )
    await exchange.connect()
    logger.info("Kết nối Binance MAINNET (chỉ đọc dữ liệu lịch sử)")

    try:
        start_dt = datetime.strptime(START_DATE, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end_dt   = datetime.strptime(END_DATE,   "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end_dt   = end_dt.replace(hour=23, minute=59, second=59)
        start_ms = int(start_dt.timestamp() * 1000)
        end_ms   = int(end_dt.timestamp()   * 1000)

        tf_ms     = _timeframe_ms(TIMEFRAME)
        warmup_ms = start_ms - 300 * tf_ms

        logger.info(f"Fetching {SYMBOL} {TIMEFRAME} intraday data...")
        all_candles = await _fetch_ohlcv_range(
            exchange._exchange, SYMBOL, TIMEFRAME, warmup_ms, end_ms
        )
        logger.info(
            f"Fetched {len(all_candles)} intraday candles | "
            f"{datetime.fromtimestamp(all_candles[0][0]/1000, tz=timezone.utc).date()} → "
            f"{datetime.fromtimestamp(all_candles[-1][0]/1000, tz=timezone.utc).date()}"
        )

        d1_days    = 200 + 14 + 10
        d1_start   = warmup_ms - d1_days * 86_400_000
        logger.info(f"Fetching {SYMBOL} 1d D1 data ({d1_days} days)...")
        d1_candles = await _fetch_ohlcv_range(
            exchange._exchange, SYMBOL, "1d", d1_start, end_ms
        )
        logger.info(
            f"Fetched {len(d1_candles)} D1 candles | "
            f"{datetime.fromtimestamp(d1_candles[0][0]/1000, tz=timezone.utc).date()} → "
            f"{datetime.fromtimestamp(d1_candles[-1][0]/1000, tz=timezone.utc).date()}"
        )
    finally:
        await exchange.close()

    if len(all_candles) < 310:
        logger.error(f"Không đủ dữ liệu intraday: {len(all_candles)}")
        return

    df_base = pd.DataFrame(
        all_candles,
        columns=["timestamp", "open", "high", "low", "close", "volume"]
    )

    # ── Pre-compute indicators ONCE ───────────────────────────────────────────
    logger.info("Pre-computing ADTS indicators (1 lần cho tất cả combinations)...")
    t0 = datetime.now()
    df_computed = _precompute_adts(df_base.copy(), FIXED_PARAMS, d1_candles=d1_candles)
    elapsed = (datetime.now() - t0).total_seconds()
    logger.info(f"Pre-compute xong trong {elapsed:.1f}s")

    calib_arr   = df_computed["_calib_sideway"].to_numpy()
    valid_count = sum(1 for v in calib_arr if not (isinstance(v, float) and np.isnan(v)))
    logger.info(f"calib_sideway: {valid_count}/{len(calib_arr)} nến hợp lệ")
    if valid_count == 0:
        logger.error("Không có calib_sideway hợp lệ — kiểm tra lại dữ liệu D1")
        return

    # Diagnostic: shield pass rate
    adx_arr = df_computed["_adx"].to_numpy()
    bbw_arr = df_computed["_bbwidth"].to_numpy()
    sid_arr = df_computed["_calib_sideway"].to_numpy()
    for adx_thr in ADX_THRESHOLD_GRID:
        n = sum(1 for v in adx_arr if not np.isnan(v) and v > adx_thr)
        logger.info(f"  ADX>{adx_thr:.0f}: {n}/{len(adx_arr)} nến ({n/len(adx_arr)*100:.1f}%)")
    n_bbw = sum(1 for i, v in enumerate(bbw_arr)
                if not np.isnan(v) and not np.isnan(sid_arr[i]) and v > sid_arr[i])
    logger.info(f"  BBW>sideway: {n_bbw}/{len(bbw_arr)} nến ({n_bbw/len(bbw_arr)*100:.1f}%)")

    # ── Grid Search ───────────────────────────────────────────────────────────
    combinations = list(itertools.product(
        SL_ATR_MULT_GRID, TP1_RR_GRID,
        ADX_THRESHOLD_GRID, BBWIDTH_FACTOR_GRID
    ))
    results = []
    t_grid  = datetime.now()

    for idx, (sl_mult, tp1_rr, adx_thr, bbw_factor) in enumerate(combinations, start=1):
        params = dict(FIXED_PARAMS)
        params["sl_atr_mult"]              = sl_mult
        params["tp1_rr"]                   = tp1_rr
        params["adx_threshold"]            = adx_thr
        params["bbwidth_threshold_factor"] = bbw_factor

        try:
            summary = _run_one(df_computed, all_candles, start_ms, end_ms, params)
        except Exception as e:
            logger.warning(f"[{idx:3d}/{len(combinations)}] Error: {e}")
            summary = None

        row = {
            "sl_atr_mult": sl_mult, "tp1_rr": tp1_rr,
            "adx_threshold": adx_thr, "bbwidth_factor": bbw_factor,
        }
        if summary is None:
            row.update({
                "total_trades": 0, "win_rate": 0, "total_pnl": 0,
                "total_return_pct": 0, "profit_factor": 0,
                "sharpe_ratio": -999, "max_drawdown_pct": 0,
                "avg_holding_candles": 0, "final_balance": INITIAL_BALANCE,
            })
        else:
            row.update(summary)
            # Log mọi combination có Sharpe > 0 hoặc mỗi 48 combination
            if idx % 48 == 0 or summary["sharpe_ratio"] > 0:
                logger.info(
                    f"[{idx:3d}/{len(combinations)}] "
                    f"sl={sl_mult} tp={tp1_rr} adx={adx_thr} bbw={bbw_factor} | "
                    f"T={summary['total_trades']:3d} WR={summary['win_rate']:5.1f}% "
                    f"PnL={summary['total_pnl']:+8.2f} "
                    f"Sharpe={summary['sharpe_ratio']:7.3f} "
                    f"MDD={summary['max_drawdown_pct']:5.1f}%"
                )
        results.append(row)

    elapsed_grid = (datetime.now() - t_grid).total_seconds()
    logger.info(
        f"Grid search hoàn tất trong {elapsed_grid:.1f}s "
        f"({elapsed_grid/len(combinations)*1000:.0f}ms/combo)"
    )

    # ── Rank & Display ────────────────────────────────────────────────────────
    df_r = pd.DataFrame(results)
    df_r = df_r[df_r["total_trades"] > 0].copy()

    if df_r.empty:
        logger.error("Không có combination nào tạo ra lệnh.")
        return

    def _norm(s):
        mn, mx = s.min(), s.max()
        return (s - mn) / (mx - mn) if mx > mn else s * 0

    df_r["score"] = (
        _norm(df_r["sharpe_ratio"])      * 0.40 +
        _norm(df_r["profit_factor"])     * 0.30 +
        _norm(df_r["win_rate"])          * 0.20 -
        _norm(df_r["max_drawdown_pct"])  * 0.10
    )
    df_r = df_r.sort_values("score", ascending=False).reset_index(drop=True)
    df_r.index += 1

    print("\n" + "=" * 105)
    print(f"  ADTS Grid Search — {SYMBOL}  {START_DATE} → {END_DATE}  (top 20 / {len(df_r)})")
    print("=" * 105)
    print(f"{'#':>3} {'SL×ATR':>7} {'TP1RR':>6} {'ADX':>5} {'BBWf':>5} "
          f"{'T':>4} {'WR%':>6} {'PnL':>9} {'Ret%':>7} "
          f"{'PF':>6} {'Sharpe':>8} {'MDD%':>7} {'Score':>7}")
    print("-" * 105)

    for rank, row in df_r.head(20).iterrows():
        marker = " ◀ BEST" if rank == 1 else (" ★" if rank <= 3 else "")
        print(
            f"{rank:>3} {row['sl_atr_mult']:>7.1f} {row['tp1_rr']:>6.1f} "
            f"{row['adx_threshold']:>5.0f} {row['bbwidth_factor']:>5.2f} "
            f"{int(row['total_trades']):>4} {row['win_rate']:>6.1f} "
            f"{row['total_pnl']:>+9.2f} {row['total_return_pct']:>+7.2f}% "
            f"{row['profit_factor']:>6.2f} {row['sharpe_ratio']:>8.3f} "
            f"{row['max_drawdown_pct']:>7.1f}% {row['score']:>7.4f}{marker}"
        )

    best = df_r.iloc[0]
    print("=" * 105)
    print(f"\n🏆 Tổ hợp tối ưu:")
    print(f"   sl_atr_mult              = {best['sl_atr_mult']:.1f}")
    print(f"   tp1_rr                   = {best['tp1_rr']:.1f}")
    print(f"   adx_threshold            = {best['adx_threshold']:.0f}")
    print(f"   bbwidth_threshold_factor = {best['bbwidth_factor']:.2f}")
    print(f"   Sharpe Ratio             = {best['sharpe_ratio']:.3f}")
    print(f"   Profit Factor            = {best['profit_factor']:.2f}")
    print(f"   Win Rate                 = {best['win_rate']:.1f}%")
    print(f"   Total PnL                = ${best['total_pnl']:+.2f}")
    print(f"   Max Drawdown             = {best['max_drawdown_pct']:.1f}%")
    print(f"   Trades                   = {int(best['total_trades'])}")

    # ── Save Excel ────────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        os.makedirs("data/backtest", exist_ok=True)
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"adts_grid_search_BTCUSDT_{date_str}.xlsx"
        filepath = os.path.join("data/backtest", filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grid Search Results"

        hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        top3_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
        red_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        bold_font = Font(bold=True)

        headers = [
            "Rank", "SL×ATR", "TP1 R:R", "ADX Thr", "BBW Factor",
            "Trades", "Win Rate%", "Total PnL", "Return%",
            "Profit Factor", "Sharpe", "MDD%", "Avg Hold", "Score"
        ]
        widths = [6, 9, 9, 9, 11, 8, 11, 14, 10, 14, 10, 8, 11, 9]

        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, (rank, row) in enumerate(df_r.iterrows(), 2):
            rf = best_fill if rank == 1 else (top3_fill if rank <= 3 else None)
            vals = [
                rank, row["sl_atr_mult"], row["tp1_rr"],
                row["adx_threshold"], row["bbwidth_factor"],
                int(row["total_trades"]), row["win_rate"],
                round(row["total_pnl"], 4), row["total_return_pct"],
                row["profit_factor"], row["sharpe_ratio"],
                row["max_drawdown_pct"], row["avg_holding_candles"],
                round(row["score"], 4),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                if rf:
                    c.fill = rf
                elif ci == 8:
                    c.fill = best_fill if (v or 0) > 0 else red_fill
                elif ci == 12 and (v or 0) > 20:
                    c.fill = red_fill

        # Heatmap: Sharpe (sl_atr_mult × tp1_rr, best adx+bbw)
        ws2 = wb.create_sheet("Heatmap Sharpe")
        ws2.cell(row=1, column=1, value="SL×ATR \\ TP1 R:R").font = bold_font
        for ci, tp1 in enumerate(TP1_RR_GRID, 2):
            ws2.cell(row=1, column=ci, value=tp1).font = bold_font
        for ri, sl in enumerate(SL_ATR_MULT_GRID, 2):
            ws2.cell(row=ri, column=1, value=sl).font = bold_font
            for ci, tp1 in enumerate(TP1_RR_GRID, 2):
                sub = df_r[(df_r["sl_atr_mult"] == sl) & (df_r["tp1_rr"] == tp1)]
                val = round(float(sub["sharpe_ratio"].max()), 3) if not sub.empty else 0.0
                c = ws2.cell(row=ri, column=ci, value=val)
                if val > 0.5:
                    c.fill = best_fill
                elif val > 0:
                    c.fill = top3_fill
                elif val < 0:
                    c.fill = red_fill

        wb.save(filepath)
        print(f"\n📊 Kết quả đã lưu: {filepath}")

    except Exception as e:
        print(f"\n⚠ Lỗi xuất Excel: {e}")


if __name__ == "__main__":
    asyncio.run(main())
