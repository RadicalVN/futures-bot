import openpyxl, sys
f = sys.argv[1] if len(sys.argv) > 1 else 'data/backtest/backtest_7_BTCUSDT_5m_20260420_20260430.xlsx'
wb = openpyxl.load_workbook(f)

print('=== TONG HOP ===')
for row in wb['Tong hop'].iter_rows(values_only=True):
    if row[0]: print(f'  {str(row[0]):<32} {row[1]}')

print()
print('=== TAT CA LENH ===')
ws2 = wb['Chi tiet lenh']
wins = losses = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    pnl = row[8] or 0
    if pnl > 0: wins += 1
    else: losses += 1
    mark = '✓' if pnl > 0 else '✗'
    print(f'  {mark} #{row[0]:>2} | {row[1]} -> {row[2]} | {str(row[4]):<6} | entry={row[5]:>10.2f} exit={row[6]:>10.2f} | pnl={pnl:>9.4f} ({row[9]:>+.2f}%) | hold={row[11]:>3}n')
print(f'\n  Wins={wins} Losses={losses}')
